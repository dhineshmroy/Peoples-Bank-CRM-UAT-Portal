import streamlit as st
import pandas as pd
import io
import ast
import psycopg2
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_db_connection():
    try:
        db_url = st.secrets["postgres"]["url"]
        conn = psycopg2.connect(db_url)
        return conn
    except Exception as e:
        st.error(f"PostgreSQL Connection Failed: {e}")
        return None

def format_module_dataframe(df_raw, module_name):
    """Maps raw database rows into the exact custom columns and title-cased headers per module."""
    formatted_rows = []
    
    for _, row in df_raw.iterrows():
        extra = {}
        extra_data_val = row.get("extra_data")
        if extra_data_val:
            if isinstance(extra_data_val, dict):
                extra = extra_data_val
            elif isinstance(extra_data_val, str):
                try:
                    extra = ast.literal_eval(extra_data_val)
                except Exception:
                    extra = {}

        if module_name == "GRG_CRM_Cardless_Bill_Payment":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Biller Name / Category": extra.get("biller_name", ""),
                "Consumer / Acc / Ref No": extra.get("consumer_acc", ""),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "Bill Txn Amount (LKR)": extra.get("bill_amount", 0.0),
                "Service Charge": extra.get("service_charge", 0.0),
                "Actual Paid Txn Amount (LKR)": extra.get("actual_paid", 0.0),
                "FE Status": row.get("fe_status", ""),
                "Biller & SV Status": extra.get("biller_sv_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name == "GRG_CRM_Cardless_Cash_Deposit":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Account Number": extra.get("account_number", ""),
                "Before Txn Balance (LKR)": extra.get("before_balance", 0.0),
                "Txn Amount (LKR)": extra.get("txn_amount", 0.0),
                "After Txn Balance (LKR)": extra.get("after_balance", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name == "GRG_CRM_Cardbased_Bill_Payment":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Bill Number": extra.get("bill_number", ""),
                "Card Number": extra.get("card_number", ""),
                "Account Number": extra.get("account_number", ""),
                "Card Type": extra.get("card_type", ""),
                "Before Txn Balance (LKR)": extra.get("before_balance", 0.0),
                "Txn Amount (LKR)": extra.get("txn_amount", 0.0),
                "Service Charge": extra.get("service_charge", 0.0),
                "Actual Paid Amount (LKR)": extra.get("actual_paid", 0.0),
                "After Txn Balance (LKR)": extra.get("after_balance", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name == "GRG_CRM_Cardbased_Cash_Deposit":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Account Number": extra.get("account_number", ""),
                "Card Type": extra.get("card_type", ""),
                "Before Txn Balance (LKR)": extra.get("before_balance", 0.0),
                "Txn Amount (LKR)": extra.get("txn_amount", 0.0),
                "After Txn Balance (LKR)": extra.get("after_balance", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name == "GRG_CRM_Cardbased_Cash_Withdraw":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Card Number": extra.get("card_number", ""),
                "Account Number": extra.get("account_number", ""),
                "Card Type": extra.get("card_type", ""),
                "Before Txn Balance (LKR)": extra.get("before_balance", 0.0),
                "Txn Amount (LKR)": extra.get("txn_amount", 0.0),
                "Service Charge": extra.get("service_charge", 0.0),
                "Actual Txn Amount (LKR)": extra.get("actual_txn", 0.0),
                "After Txn Balance (LKR)": extra.get("after_balance", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name == "GRG_CRM_Cardbased_Fund_transfer":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Card Number": extra.get("card_number", ""),
                "Card Type": extra.get("card_type", ""),
                "From Account Number": extra.get("from_acc", ""),
                "To Account Number": extra.get("to_acc", ""),
                "Before Balance From Acc (LKR)": extra.get("before_bal_from", 0.0),
                "Before Balance To Acc (LKR)": extra.get("before_bal_to", 0.0),
                "Txn Amount (LKR)": extra.get("txn_amount", 0.0),
                "After Balance From Acc (LKR)": extra.get("after_bal_from", 0.0),
                "After Balance To Acc (LKR)": extra.get("after_bal_to", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name in ["GRG_CRM_Cardbased_SLIC_Bill_Pay", "GRG_CRM_Cardless_SLIC_Bill_Paym"]:
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Bill Number": extra.get("bill_number", ""),
                "Card / Biller Type": extra.get("card_type", ""),
                "Txn Amount (LKR)": extra.get("txn_amount", 0.0),
                "Service Charge": extra.get("service_charge", 0.0),
                "Actual Paid Amount (LKR)": extra.get("actual_paid", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        elif module_name == "GRG_CRM_Cardbased_Mini_Statemen":
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "Card Number": extra.get("card_number", ""),
                "Account Number": extra.get("account_number", ""),
                "Card Type": extra.get("card_type", ""),
                "Before Txn Balance (LKR)": extra.get("before_balance", 0.0),
                "Service Charge": extra.get("service_charge", 0.0),
                "Actual Txn Amount (LKR)": extra.get("actual_txn", 0.0),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
        else:
            formatted_rows.append({
                "TC ID": row.get("tc_id", ""),
                "Test Description": row.get("test_description", ""),
                "RRN (Retrieval Ref)": row.get("rrn", ""),
                "STAN / UTANO": row.get("stan_utano", ""),
                "FE Status": row.get("fe_status", ""),
                "Core Banking (SIBS) Status": row.get("sibs_status", ""),
                "Overall Test Status": row.get("overall_status", ""),
                "Expected Result / Remarks": row.get("tester_remarks", "")
            })
            
    return pd.DataFrame(formatted_rows)

def render_test_execution_page():
    st.title("💳 GRG CRM - Test Execution & Cash Loading Management")
    
    tab_exec, tab_cash, tab_export = st.tabs(["📝 Test Execution", "💵 Cash Loading & Receipts", "📊 Finance Export"])

    modules = [
        "GRG_CRM_Cardless_Bill_Payment",
        "GRG_CRM_Cardless_Cash_Deposit",
        "GRG_CRM_Cardbased_Bill_Payment",
        "GRG_CRM_Cardbased_Cash_Deposit",
        "GRG_CRM_Cardbased_Cash_Withdraw",
        "GRG_CRM_Cardbased_Fund_transfer",
        "GRG_CRM_Cardbased_SLIC_Bill_Pay",
        "GRG_CRM_Cardless_SLIC_Bill_Paym",
        "GRG_CRM_Cardbased_Mini_Statemen"
    ]

    # -------------------------------------------------------------------------
    # TAB 1: TEST EXECUTION WITH UNIQUE UPSERT
    # -------------------------------------------------------------------------
    with tab_exec:
        st.subheader("Execute Pre-Built Test Cases (Dynamic Module Layouts)")
        selected_module = st.selectbox("Select Test Module / Feature", modules, key="exec_mod")

        conn = get_db_connection()
        tc_list = []
        tc_data_dict = {}
        if conn:
            try:
                tc_df = pd.read_sql(f"SELECT tc_id, test_description, biller_category, expected_remarks FROM uat_test_cases WHERE module_name = '{selected_module}'", conn)
                conn.close()
                if not tc_df.empty:
                    tc_list = tc_df['tc_id'].tolist()
                    for _, row in tc_df.iterrows():
                        tc_data_dict[row['tc_id']] = {
                            "description": row.get('test_description', ''),
                            "biller": row.get('biller_category', ''),
                            "remarks": row.get('expected_remarks', '')
                        }
            except Exception:
                pass

        if not tc_list:
            st.warning("No pre-built test cases found in database for this module.")
            selected_tc = st.text_input("Test Case ID (Manual)", value="TC_001")
            pre_desc, pre_biller, pre_remarks = "", "", ""
        else:
            selected_tc = st.selectbox("Select Test Case ID", tc_list)
            details = tc_data_dict.get(selected_tc, {})
            pre_desc = details.get("description", "")
            pre_biller = details.get("biller", "")
            pre_remarks = details.get("remarks", "")
            st.info(f"**Test Description:** {pre_desc}")

        col1, col2 = st.columns(2)
        with col1:
            def update_rrn():
                stan_val = st.session_state.get("stan_input", "")
                if len(stan_val) >= 12:
                    st.session_state.rrn_input = stan_val[-12:]
                else:
                    st.session_state.rrn_input = stan_val

            stan = st.text_input("STAN / UTANO", placeholder="Enter STAN / UTANO here...", key="stan_input", on_change=update_rrn)
            
            if "rrn_input" not in st.session_state:
                st.session_state.rrn_input = ""

            rrn = st.text_input("RRN (Retrieval Reference Number - Auto)", key="rrn_input")

        form_data = {}

        if selected_module == "GRG_CRM_Cardless_Bill_Payment":
            with col2:
                form_data["biller_name"] = st.text_input("Biller Name / Category", value=pre_biller)
                form_data["consumer_acc"] = st.text_input("Consumer / Acc / Ref No")
            c3, c4 = st.columns(2)
            with c3:
                bill_amount = st.number_input("Bill Txn Amount (LKR)", value=0.00, format="%.2f", key="cardless_bill_amt")
                service_charge = st.number_input("Service Charge (LKR)", value=0.00, format="%.2f", key="cardless_serv_chg")
            with c4:
                actual_paid = bill_amount + service_charge
                form_data["bill_amount"] = bill_amount
                form_data["service_charge"] = service_charge
                form_data["actual_paid"] = actual_paid
                st.markdown(f"### **Actual Paid Txn Amount (LKR) [Auto]:** `LKR {actual_paid:,.2f}`")
                biller_sv_status = st.selectbox("Biller & SV Status", ["UPDATED", "PENDING", "FAILED"])
                form_data["biller_sv_status"] = biller_sv_status

        elif selected_module == "GRG_CRM_Cardless_Cash_Deposit":
            with col2:
                form_data["account_number"] = st.text_input("Account Number")
            c3, c4 = st.columns(2)
            with c3:
                form_data["before_balance"] = st.number_input("Before Txn Balance (LKR)", value=0.00, format="%.2f")
                form_data["txn_amount"] = st.number_input("Txn Amount (LKR)", value=0.00, format="%.2f")
            with c4:
                form_data["after_balance"] = st.number_input("After Txn Balance (LKR)", value=0.00, format="%.2f")

        elif selected_module == "GRG_CRM_Cardbased_Bill_Payment":
            with col2:
                form_data["bill_number"] = st.text_input("Bill Number")
                form_data["card_number"] = st.text_input("Card Number")
            c3, c4 = st.columns(2)
            with c3:
                form_data["account_number"] = st.text_input("Account Number")
                form_data["card_type"] = st.selectbox("Card Type", ["VISA", "MASTER", "AMEX", "RUPAY"])
                form_data["before_balance"] = st.number_input("Before Txn Balance (LKR)", value=0.00, format="%.2f")
                txn_amt = st.number_input("Txn Amount (LKR)", value=0.00, format="%.2f", key="cb_bill_amt")
            with c4:
                serv_chg = st.number_input("Service Charge (LKR)", value=0.00, format="%.2f", key="cb_bill_serv")
                actual_paid = txn_amt + serv_chg
                form_data["txn_amount"] = txn_amt
                form_data["service_charge"] = serv_chg
                form_data["actual_paid"] = actual_paid
                st.markdown(f"### **Actual Paid Amount (LKR) [Auto]:** `LKR {actual_paid:,.2f}`")
                form_data["after_balance"] = st.number_input("After Txn Balance (LKR)", value=0.00, format="%.2f")

        elif selected_module == "GRG_CRM_Cardbased_Cash_Deposit":
            with col2:
                form_data["account_number"] = st.text_input("Account Number")
                form_data["card_type"] = st.selectbox("Card Type", ["VISA", "MASTER", "AMEX"])
            c3, c4 = st.columns(2)
            with c3:
                form_data["before_balance"] = st.number_input("Before Txn Balance (LKR)", value=0.00, format="%.2f")
                form_data["txn_amount"] = st.number_input("Txn Amount (LKR)", value=0.00, format="%.2f")
            with c4:
                form_data["after_balance"] = st.number_input("After Txn Balance (LKR)", value=0.00, format="%.2f")

        elif selected_module == "GRG_CRM_Cardbased_Cash_Withdraw":
            with col2:
                form_data["card_number"] = st.text_input("Card Number")
                form_data["account_number"] = st.text_input("Account Number")
            c3, c4 = st.columns(2)
            with c3:
                form_data["card_type"] = st.selectbox("Card Type", ["VISA", "MASTER", "AMEX"])
                form_data["before_balance"] = st.number_input("Before Txn Balance (LKR)", value=0.00, format="%.2f")
                txn_amt = st.number_input("Txn Amount (LKR)", value=0.00, format="%.2f", key="cb_wd_amt")
            with c4:
                serv_chg = st.number_input("Service Charge (LKR)", value=0.00, format="%.2f", key="cb_wd_serv")
                actual_txn = txn_amt + serv_chg
                form_data["txn_amount"] = txn_amt
                form_data["service_charge"] = serv_chg
                form_data["actual_txn"] = actual_txn
                st.markdown(f"### **Actual Txn Amount (LKR) [Auto]:** `LKR {actual_txn:,.2f}`")
                form_data["after_balance"] = st.number_input("After Txn Balance (LKR)", value=0.00, format="%.2f")

        elif selected_module == "GRG_CRM_Cardbased_Fund_transfer":
            with col2:
                form_data["card_number"] = st.text_input("Card Number")
                form_data["card_type"] = st.selectbox("Card Type", ["VISA", "MASTER"])
            c3, c4 = st.columns(2)
            with c3:
                form_data["from_acc"] = st.text_input("From Account Number")
                form_data["to_acc"] = st.text_input("To Account Number")
                form_data["before_bal_from"] = st.number_input("Before Txn Balance (LKR) - From Acc", value=0.00, format="%.2f")
                form_data["before_bal_to"] = st.number_input("Before Txn Balance (LKR) - To Acc", value=0.00, format="%.2f")
            with c4:
                form_data["txn_amount"] = st.number_input("Txn Amount (LKR)", value=0.00, format="%.2f")
                form_data["after_bal_from"] = st.number_input("After Txn Balance (LKR) - From Acc", value=0.00, format="%.2f")
                form_data["after_bal_to"] = st.number_input("After Txn Balance (LKR) - To Acc", value=0.00, format="%.2f")

        elif selected_module in ["GRG_CRM_Cardbased_SLIC_Bill_Pay", "GRG_CRM_Cardless_SLIC_Bill_Paym"]:
            with col2:
                form_data["bill_number"] = st.text_input("Bill Number")
                if "Cardbased" in selected_module:
                    form_data["card_number"] = st.text_input("Card Number")
                    form_data["account_number"] = st.text_input("Account Number")
                form_data["card_type"] = st.selectbox("Card Type / Biller Type", ["SLIC", "INSURANCE", "VISA", "MASTER"])
            c3, c4 = st.columns(2)
            with c3:
                txn_amt = st.number_input("Txn Amount (LKR)", value=0.00, format="%.2f", key="slic_amt")
                serv_chg = st.number_input("Service Charge (LKR)", value=0.00, format="%.2f", key="slic_serv")
            with c4:
                actual_paid = txn_amt + serv_chg
                form_data["txn_amount"] = txn_amt
                form_data["service_charge"] = serv_chg
                form_data["actual_paid"] = actual_paid
                st.markdown(f"### **Actual Paid Amount (LKR) [Auto]:** `LKR {actual_paid:,.2f}`")

        elif selected_module == "GRG_CRM_Cardbased_Mini_Statemen":
            with col2:
                form_data["card_number"] = st.text_input("Card Number")
                form_data["account_number"] = st.text_input("Account Number")
            c3, c4 = st.columns(2)
            with c3:
                form_data["card_type"] = st.selectbox("Card Type", ["VISA", "MASTER"])
                form_data["before_balance"] = st.number_input("Before Txn Balance (LKR)", value=0.00, format="%.2f")
            with c4:
                serv_chg = st.number_input("Service Charge (LKR)", value=0.00, format="%.2f", key="mini_serv")
                form_data["service_charge"] = serv_chg
                form_data["actual_txn"] = serv_chg
                st.markdown(f"### **Actual Txn Amount (LKR):** `LKR {serv_chg:,.2f}`")

        # Common Status Controls
        st.markdown("---")
        s_col1, s_col2 = st.columns(2)
        with s_col1:
            fe_status = st.selectbox("Front-End (FE) Status", ["SUCCESS", "FAILED"])
        with s_col2:
            sibs_status = st.selectbox("Core Banking (SIBS) Status", ["UPDATED (SIBS)", "PENDING", "FAILED"])

        overall_status = "PASS" if fe_status == "SUCCESS" and sibs_status == "UPDATED (SIBS)" else "FAIL"
        st.write(f"**Overall Test Status (Auto):** `{overall_status}`")

        remarks = st.text_area("Tester Remarks / Notes", value=pre_remarks)
        
        if st.button("💾 Save Test Execution Record", type="primary"):
            conn = get_db_connection()
            if conn:
                try:
                    cur = conn.cursor()
                    # Ensure unique latest record per test case
                    cur.execute("""
                        DELETE FROM uat_test_executions 
                        WHERE module_name = %s AND tc_id = %s;
                    """, (selected_module, selected_tc))

                    try:
                        cur.execute("""
                            INSERT INTO uat_test_executions 
                            (module_name, tc_id, test_description, rrn, stan_utano, fe_status, sibs_status, overall_status, tester_remarks, executed_by, extra_data)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (selected_module, selected_tc, pre_desc, rrn, stan, fe_status, sibs_status, overall_status, remarks, st.session_state.get("logged_user", "TESTER"), str(form_data)))
                    except Exception:
                        conn.rollback()
                        cur.execute("""
                            INSERT INTO uat_test_executions 
                            (module_name, tc_id, test_description, rrn, stan_utano, fe_status, sibs_status, overall_status, tester_remarks, executed_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (selected_module, selected_tc, pre_desc, rrn, stan, fe_status, sibs_status, overall_status, remarks, st.session_state.get("logged_user", "TESTER")))
                    
                    conn.commit()
                    cur.close()
                    conn.close()
                    st.success(f"Successfully saved latest update for **{selected_tc}** under **{selected_module}**!")
                except Exception as e:
                    st.error(f"Error saving execution record: {e}")

    # -------------------------------------------------------------------------
    # TAB 2: CASH LOADING & UNLOADING MANAGEMENT (FULL CODE)
    # -------------------------------------------------------------------------
    with tab_cash:
        st.subheader("💵 Terminal Cash Loading & Unloading Tracker")
        st.markdown("Record individual cash loading sessions (1st to 10th) with amounts and receipts, followed by the final Unloading session receipts.")
        
        # Sub-tabs for Loading vs Unloading
        cash_sub_tab1, cash_sub_tab2 = st.tabs(["📥 Cash Loading Sessions (1 to 10)", "📤 Final Unloading Session & Receipts"])

        # -----------------------------------------------------------------
        # SUB-TAB 1: CASH LOADING SESSIONS
        # -----------------------------------------------------------------
        with cash_sub_tab1:
            st.markdown("### Record Cash Loading (1st up to 10th)")
            with st.form("cash_loading_form_unique_v2"):
                c_col1, c_col2 = st.columns(2)
                with c_col1:
                    terminal_id_load = st.text_input("Terminal ID", value="S169RB02", key="load_term_id_v2")
                    report_date_load = st.date_input("Report Date", value=datetime.today(), key="load_rep_date_v2")
                    loading_session = st.selectbox(
                        "Select Loading Session", 
                        [
                            "1st Cash Loading", "2nd Cash Loading", "3rd Cash Loading", 
                            "4th Cash Loading", "5th Cash Loading", "6th Cash Loading", 
                            "7th Cash Loading", "8th Cash Loading", "9th Cash Loading", 
                            "10th Cash Loading"
                        ],
                        key="load_session_type_v2"
                    )
                with c_col2:
                    load_time = st.time_input("Loading Action Time", key="load_time_val_v2")
                    loading_total = st.number_input("Loading Session Total Amount (LKR)", value=0.00, min_value=0.00, format="%.2f", key="load_amt_val_v2")

                st.markdown("---")
                st.markdown(f"#### 📄 Mandatory SOP & HOST Receipts for `{loading_session}`")
                sop_file = st.file_uploader(f"Upload SOP Receipt for {loading_session} (.pdf, .png, .jpg)", type=["pdf", "png", "jpg"], key="load_sop_file_v2")
                host_file = st.file_uploader(f"Upload HOST Receipt for {loading_session} (.pdf, .png, .jpg)", type=["pdf", "png", "jpg"], key="load_host_file_v2")

                if st.form_submit_button(f"💾 Save {loading_session} Entry", type="primary"):
                    sop_name = sop_file.name if sop_file else "None"
                    host_name = host_file.name if host_file else "None"
                    conn = get_db_connection()
                    if conn:
                        try:
                            cur = conn.cursor()
                            cur.execute("""
                                CREATE TABLE IF NOT EXISTS terminal_cash_logs (
                                    id SERIAL PRIMARY KEY,
                                    terminal_id VARCHAR(50),
                                    report_date DATE,
                                    loading_session VARCHAR(50),
                                    load_time VARCHAR(20),
                                    loading_total NUMERIC(15,2),
                                    sop_receipt_path VARCHAR(255),
                                    host_receipt_path VARCHAR(255),
                                    logged_by VARCHAR(50),
                                    logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                    CONSTRAINT unique_terminal_session UNIQUE (terminal_id, report_date, loading_session)
                                );
                            """)
                            
                            cur.execute("""
                                INSERT INTO terminal_cash_logs 
                                (terminal_id, report_date, loading_session, load_time, loading_total, sop_receipt_path, host_receipt_path, logged_by)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                                ON CONFLICT (terminal_id, report_date, loading_session)
                                DO UPDATE SET 
                                    load_time = EXCLUDED.load_time,
                                    loading_total = EXCLUDED.loading_total,
                                    sop_receipt_path = EXCLUDED.sop_receipt_path,
                                    host_receipt_path = EXCLUDED.host_receipt_path,
                                    logged_at = CURRENT_TIMESTAMP;
                            """, (terminal_id_load, report_date_load, loading_session, str(load_time), loading_total, sop_name, host_name, st.session_state.get("logged_user", "TESTER")))
                            
                            conn.commit()
                            cur.close()
                            conn.close()
                            st.success(f"Successfully saved **{loading_session}** (Amount: LKR {loading_total:,.2f}) with its SOP & HOST receipts!")
                        except Exception as e:
                            st.error(f"Error saving loading log: {e}")

        # -----------------------------------------------------------------
        # SUB-TAB 2: FINAL UNLOADING SESSION
        # -----------------------------------------------------------------
        with cash_sub_tab2:
            st.markdown("### Record Final Unloading Receipts")
            st.info("⚠️ Complete all loading sessions first. Before final unloading, select the action time and upload both final unloading receipts.")
            
            with st.form("cash_unloading_form_unique_v2"):
                u_col1, u_col2 = st.columns(2)
                with u_col1:
                    unloading_term_id = st.text_input("Terminal ID", value="S169RB02", key="unload_term_id_v2")
                    unloading_date = st.date_input("Report Date", value=datetime.today(), key="unload_rep_date_v2")
                with u_col2:
                    unload_time = st.time_input("Unloading Action Time", key="unload_time_val_v2")

                st.markdown("---")
                st.markdown("#### 📄 Mandatory Final Unloading Receipts (SOP & HOST)")
                unload_sop = st.file_uploader("Upload Final SOP Receipt for Unloading (.pdf, .png, .jpg)", type=["pdf", "png", "jpg"], key="unload_sop_file_v2")
                unload_host = st.file_uploader("Upload Final HOST Receipt for Unloading (.pdf, .png, .jpg)", type=["pdf", "png", "jpg"], key="unload_host_file_v2")

                if st.form_submit_button("💾 Save Final Unloading Receipts", type="primary"):
                    sop_name = unload_sop.name if unload_sop else "None"
                    host_name = unload_host.name if unload_host else "None"
                    conn = get_db_connection()
                    if conn:
                        try:
                            cur = conn.cursor()
                            cur.execute("""
                                CREATE TABLE IF NOT EXISTS terminal_cash_logs (
                                    id SERIAL PRIMARY KEY,
                                    terminal_id VARCHAR(50),
                                    report_date DATE,
                                    loading_session VARCHAR(50),
                                    load_time VARCHAR(20),
                                    loading_total NUMERIC(15,2),
                                    sop_receipt_path VARCHAR(255),
                                    host_receipt_path VARCHAR(255),
                                    logged_by VARCHAR(50),
                                    logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                    CONSTRAINT unique_terminal_session UNIQUE (terminal_id, report_date, loading_session)
                                );
                            """)
                            
                            cur.execute("""
                                INSERT INTO terminal_cash_logs 
                                (terminal_id, report_date, loading_session, load_time, loading_total, sop_receipt_path, host_receipt_path, logged_by)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                                ON CONFLICT (terminal_id, report_date, loading_session)
                                DO UPDATE SET 
                                    load_time = EXCLUDED.load_time,
                                    loading_total = EXCLUDED.loading_total,
                                    sop_receipt_path = EXCLUDED.sop_receipt_path,
                                    host_receipt_path = EXCLUDED.host_receipt_path,
                                    logged_at = CURRENT_TIMESTAMP;
                            """, (unloading_term_id, unloading_date, "Unloading Session", str(unload_time), 0.00, sop_name, host_name, st.session_state.get("logged_user", "TESTER")))
                            
                            conn.commit()
                            cur.close()
                            conn.close()
                            st.success("Successfully saved **Unloading Session** along with both final SOP & HOST receipts!")
                        except Exception as e:
                            st.error(f"Error saving unloading log: {e}")

    # -------------------------------------------------------------------------
    # TAB 3: FINANCE EXPORT & INTERACTIVE TABLE VIEWER
    # -------------------------------------------------------------------------
    with tab_export:
        st.subheader("📊 Finance Export & Interactive Table Viewer")
        st.markdown("View, filter, sort, and download complete multi-tab execution reports styled professionally.")

        view_options = ["🌐 All Modules (Combined Master View)", "💵 Cash Loading & Unloading Report"] + modules
        selected_view_option = st.selectbox("Select Module to View/Sort", view_options, key="view_mod")
        
        conn = get_db_connection()
        df_raw = pd.DataFrame()
        if conn:
            try:
                if selected_view_option == "🌐 All Modules (Combined Master View)":
                    df_raw = pd.read_sql("SELECT * FROM uat_test_executions", conn)
                elif selected_view_option == "💵 Cash Loading & Unloading Report":
                    df_raw = pd.read_sql("SELECT * FROM terminal_cash_logs ORDER BY report_date DESC", conn)
                else:
                    df_raw = pd.read_sql(f"SELECT * FROM uat_test_executions WHERE module_name = '{selected_view_option}'", conn)
                conn.close()
            except Exception:
                pass

        if not df_raw.empty:
            if selected_view_option == "🌐 All Modules (Combined Master View)":
                df_view = df_raw
            elif selected_view_option == "💵 Cash Loading & Unloading Report":
                df_view = df_raw
            else:
                df_view = format_module_dataframe(df_raw, selected_view_option)

            st.markdown(f"### 📋 Records for `{selected_view_option}`")
            
            sort_col = st.selectbox("Sort By Column", df_view.columns.tolist(), key="sort_col")
            sort_order = st.radio("Sort Order", ["Ascending", "Descending"], horizontal=True, key="sort_ord")
            ascending_bool = True if sort_order == "Ascending" else False
            
            df_sorted = df_view.sort_values(by=sort_col, ascending=ascending_bool)
            st.dataframe(df_sorted, use_container_width=True)
        else:
            st.info(f"No records found for `{selected_view_option}` yet.")

        st.markdown("---")
        st.subheader("📥 Download Styled Multi-Tab Finance Workbook")
        
        if st.button("📥 Generate & Download Colorful Excel Report"):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                conn = get_db_connection()
                if conn:
                    try:
                        # 1. Cash Loading & Unloading Summary Sheet (Formatted exactly as requested)
                        try:
                            df_cash = pd.read_sql("SELECT * FROM terminal_cash_logs ORDER BY report_date DESC LIMIT 1", conn)
                            if not df_cash.empty:
                                c_row = df_cash.iloc[0]
                                term_id_val = c_row.get("terminal_id", "S169RB02")
                                date_val = str(c_row.get("report_date", datetime.today().strftime('%Y-%m-%d')))
                            else:
                                term_id_val, date_val = "S169RB02", datetime.today().strftime('%Y-%m-%d')
                            
                            # Fetch totals per session
                            df_all_cash = pd.read_sql("SELECT loading_session, load_time, loading_total FROM terminal_cash_logs", conn)
                            session_dict = {}
                            total_loaded = 0.0
                            unloading_total = 0.0
                            for _, r in df_all_cash.iterrows():
                                s_name = r["loading_session"]
                                session_dict[f"{s_name}_time"] = r["load_time"]
                                session_dict[f"{s_name}_total"] = float(r["loading_total"] or 0)
                                if "Loading" in s_name:
                                    total_loaded += float(r["loading_total"] or 0)
                                elif "Unloading" in s_name:
                                    unloading_total += float(r["loading_total"] or 0)
                        except Exception:
                            term_id_val = "S169RB02"
                            date_val = datetime.today().strftime('%Y-%m-%d')
                            session_dict = {}
                            total_loaded = 0.0
                            unloading_total = 0.0

                        cash_report_data = [
                            ["Terminal ID", term_id_val],
                            ["Date", date_val],
                            ["1st Cash Loading Time", session_dict.get("1st Cash Loading_time", "-")],
                            ["1st Loading Total", session_dict.get("1st Cash Loading_total", 0.0)],
                            ["2nd Cash Loading Time", session_dict.get("2nd Cash Loading_time", "-")],
                            ["2nd Loading Total", session_dict.get("2nd Cash Loading_total", 0.0)],
                            ["3rd Cash Loading Time", session_dict.get("3rd Cash Loading_time", "-")],
                            ["3rd Loading Total", session_dict.get("3rd Cash Loading_total", 0.0)],
                            ["Total Cash Loaded", total_loaded],
                            ["Unloading Time", session_dict.get("Unloading Session_time", "-")],
                            ["Unloading Total", unloading_total],
                            ["Net Cash Balance / Variance", total_loaded - unloading_total]
                        ]
                        df_cash_report = pd.DataFrame(cash_report_data, columns=["Report Information", "Details"])
                        df_cash_report.to_excel(writer, sheet_name="Cash_Loading_Unloading", index=False, startrow=2)

                        # 2. Master Tab
                        df_all_raw = pd.read_sql("SELECT * FROM uat_test_executions", conn)
                        df_all_raw.to_excel(writer, sheet_name="All_Modules_Master", index=False, startrow=2)

                        # 3. Individual Module Tabs
                        for mod_name in modules:
                            df_mod_raw = pd.read_sql(f"SELECT * FROM uat_test_executions WHERE module_name = '{mod_name}'", conn)
                            df_formatted = format_module_dataframe(df_mod_raw, mod_name)
                            df_formatted.to_excel(writer, sheet_name=mod_name[:31], index=False, startrow=2)
                        conn.close()
                    except Exception as e:
                        st.error(f"Error compiling export: {e}")
            
            # Post-process workbook with openpyxl for exact custom formatting
            output.seek(0)
            import openpyxl
            wb = openpyxl.load_workbook(output)
            
            navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.views.sheetView[0].showGridLines = True
                
                max_col = ws.max_column
                
                # Custom Title for Cash Report vs Test Execution Reports
                if sheet_name == "Cash_Loading_Unloading":
                    title_text = "CASH LOADING & UNLOADING REPORT"
                else:
                    title_text = f"UAT Test Execution Report - {sheet_name}"

                # Merge Row 1 for the title block
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                title_cell = ws.cell(row=1, column=1, value=title_text)
                title_cell.font = title_font
                title_cell.fill = navy_fill
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 30
                
                # Style header row (Row 3)
                ws.row_dimensions[3].height = 25
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.fill = navy_fill
                    cell.font = white_bold_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
                
                # Apply center alignment and borders to all data cells
                for row in range(4, ws.max_row + 1):
                    ws.row_dimensions[row].height = 20
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Auto-fit column widths
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

            final_output = io.BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            st.download_button(
                label="⬇️ Download Professional Excel Report (.xlsx)",
                data=final_output,
                file_name=f"PeoplesBank_CRM_Finance_Master_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    render_test_execution_page()