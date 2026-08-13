import streamlit as st
import pandas as pd
from datetime import datetime, date
import io
import oracledb
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2


# ---------------------------------------------------------
# POSTGRESQL / SUPABASE DATABASE CONNECTION
# ---------------------------------------------------------
def get_db_connection():
    try:
        db_url = st.secrets["postgres"]["url"]
        conn = psycopg2.connect(db_url)
        return conn
    except Exception as e:
        st.error(f"PostgreSQL Connection Failed: {e}")
        return None

# Page Configuration
st.set_page_config(
    page_title="People's Bank | GRG CRM UAT Portal",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded"
)

UPLOAD_DIR = "uploads"
os.makedirs(os.path.join(UPLOAD_DIR, "receipts"), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_DIR, "photos"), exist_ok=True)

def clean_val(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ["nan", "none", "", "nat"]:
        return None
    return s

def load_data_from_db():
    conn = get_db_connection()
    if not conn:
        return pd.DataFrame()
    
    try:
        query = """
            SELECT 
                "TC_ID" as "TC ID", 
                "CATEGORY" as "Category",
                "MODULE_NAME" as "Module Name", 
                "TEST_AREA" as "Test Area", 
                "TEST_CASE_DESCRIPTION" as "Test Case Description", 
                "PRE_CONDITIONS" as "Pre-Conditions",
                "TEST_STEPS" as "Test Steps", 
                "EXPECTED_RESULT" as "Expected Result", 
                "PATH_TYPE" as "Path Type", 
                "ACTUAL_RESULT" as "Actual Result", 
                "RRN" as "RRN",
                "UTANO" as "Utano", 
                "STATUS" as "Status", 
                "FE" as "FE", 
                "SIBS" as "SIBS",
                "REMARKS" as "Remarks", 
                "EXECUTED_BY" as "Executed By", 
                "EXECUTED_DATE" as "Executed Date", 
                "RECEIPT_PATH" as "Receipt_Path",
                "PHOTO_PATH" as "Photo_Path",
                "SEVERITY" as "Severity",
                "PRIORITY" as "Priority",
                "DEFECT_STATUS" as "Defect Status",
                "ASSIGNED_TO" as "Assigned To",
                "TARGET_DATE" as "Target Date",
                "ROOT_CAUSE" as "Root Cause",
                "DEFECT_DESCRIPTION" as "Defect Description"
            FROM uat_test_cases_v2
        """
        df = pd.read_sql(query, con=conn)
        conn.close()
        df = df.fillna("")
        return df
    except Exception as e:
        if conn:
            conn.close()
        st.error(f"Error loading data from PostgreSQL: {e}")
        return pd.DataFrame()

def derive_rrn(utano_val):
    u_str = str(utano_val).strip()
    if not u_str or u_str.lower() in ["nan", "none", "", "nat"]:
        return "N/A"
    if u_str.endswith(".0"):
        u_str = u_str[:-2]
    if len(u_str) > 6:
        return u_str[6:]
    return u_str

def delete_test_case_from_db(tc_id, module_name):
    conn = get_db_connection()
    if not conn:
        return False
    cursor = conn.cursor()
    try:
        cursor.execute("""
            DELETE FROM uat_test_cases_v2 
            WHERE "TC_ID" = %s AND "MODULE_NAME" = %s
        """, (clean_val(tc_id), clean_val(module_name)))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Failed to delete test case from database: {e}")
        cursor.close()
        conn.close()
        return False

def safe_basename(path_val):
    p_clean = clean_val(path_val)
    if not p_clean:
        return ""
    try:
        return os.path.basename(p_clean)
    except Exception:
        return ""

def save_test_case_to_db(tc_id, module_name, status, actual_result, fe, sibs, utano, remarks, executed_by, executed_date, receipt_path, photo_path, defect_desc=""):
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    
    c_actual = clean_val(actual_result)
    c_fe = clean_val(fe)
    c_sibs = clean_val(sibs)
    c_utano = clean_val(utano)
    
    rrn_val = derive_rrn(c_utano)
    c_rrn = None if rrn_val == "N/A" else rrn_val
    
    c_remarks = clean_val(remarks)
    c_exec_by = clean_val(executed_by)
    c_exec_date = clean_val(executed_date)
    c_rec_path = clean_val(receipt_path)
    c_pho_path = clean_val(photo_path)
    c_def_desc = clean_val(defect_desc)
    
    try:
        if c_def_desc:
            cursor.execute("""
                UPDATE uat_test_cases_v2 
                SET "STATUS" = %s, "ACTUAL_RESULT" = %s, "FE" = %s, "SIBS" = %s, "UTANO" = %s, "RRN" = %s, 
                    "REMARKS" = %s, "EXECUTED_BY" = %s, "EXECUTED_DATE" = %s, "RECEIPT_PATH" = %s, "PHOTO_PATH" = %s,
                    "DEFECT_DESCRIPTION" = %s, "DEFECT_STATUS" = 'Open'
                WHERE "TC_ID" = %s AND "MODULE_NAME" = %s
            """, (status, c_actual, c_fe, c_sibs, c_utano, c_rrn, c_remarks, c_exec_by, c_exec_date, c_rec_path, c_pho_path, c_def_desc, tc_id, module_name))
        else:
            cursor.execute("""
                UPDATE uat_test_cases_v2 
                SET "STATUS" = %s, "ACTUAL_RESULT" = %s, "FE" = %s, "SIBS" = %s, "UTANO" = %s, "RRN" = %s, 
                    "REMARKS" = %s, "EXECUTED_BY" = %s, "EXECUTED_DATE" = %s, "RECEIPT_PATH" = %s, "PHOTO_PATH" = %s
                WHERE "TC_ID" = %s AND "MODULE_NAME" = %s
            """, (status, c_actual, c_fe, c_sibs, c_utano, c_rrn, c_remarks, c_exec_by, c_exec_date, c_rec_path, c_pho_path, tc_id, module_name))
        conn.commit()
    except Exception as e:
        st.error(f"Database error updating test case: {e}")
    cursor.close()
    conn.close()

def admin_update_full_defect_details(tc_id, module_name, test_steps, actual_result, executed_by, utano, fe, sibs, severity, priority, defect_status, assigned_to, target_date, root_cause, defect_desc=""):
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    
    c_steps = clean_val(test_steps)
    c_act = clean_val(actual_result)
    c_exec_by = clean_val(executed_by)
    c_utano = clean_val(utano)
    rrn_val = derive_rrn(c_utano)
    c_rrn = None if rrn_val == "N/A" else rrn_val
    c_fe = clean_val(fe)
    c_sibs = clean_val(sibs)
    c_sev = clean_val(severity)
    c_pri = clean_val(priority)
    c_def_status = clean_val(defect_status)
    c_assigned = clean_val(assigned_to)
    c_target = clean_val(target_date)
    c_root = clean_val(root_cause)
    c_def_desc = clean_val(defect_desc)

    try:
        cursor.execute("""
            UPDATE uat_test_cases_v2 
            SET "TEST_STEPS" = %s, "ACTUAL_RESULT" = %s, "EXECUTED_BY" = %s, "UTANO" = %s, "RRN" = %s, 
                "FE" = %s, "SIBS" = %s, "SEVERITY" = %s, "PRIORITY" = %s, "DEFECT_STATUS" = %s, 
                "ASSIGNED_TO" = %s, "TARGET_DATE" = %s, "ROOT_CAUSE" = %s, "DEFECT_DESCRIPTION" = %s
            WHERE "TC_ID" = %s AND "MODULE_NAME" = %s
        """, (c_steps, c_act, c_exec_by, c_utano, c_rrn, c_fe, c_sibs, c_sev, c_pri, c_def_status, 
              c_assigned, c_target, c_root, c_def_desc, tc_id, module_name))
        conn.commit()
    except Exception as e:
        st.error(f"Database error updating defect details: {e}")
    cursor.close()
    conn.close()
    
    if hasattr(st, "cache_data"):
        st.cache_data.clear()

def insert_new_test_case_to_db(tc_id, category, module_name, test_area, test_desc, pre_cond, test_steps, exp_result, path_type):
    conn = get_db_connection()
    if not conn:
        return False
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO uat_test_cases_v2 (
                "TC_ID", "CATEGORY", "MODULE_NAME", "TEST_AREA", "TEST_CASE_DESCRIPTION", 
                "PRE_CONDITIONS", "TEST_STEPS", "EXPECTED_RESULT", "PATH_TYPE", "STATUS",
                "SEVERITY", "PRIORITY", "DEFECT_STATUS", "ASSIGNED_TO"
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'PENDING', 'Medium', 'Medium', 'Open', 'Development Team')
        """, (clean_val(tc_id), clean_val(category), clean_val(module_name), clean_val(test_area), 
              clean_val(test_desc), clean_val(pre_cond), clean_val(test_steps), clean_val(exp_result), clean_val(path_type)))
        conn.commit()
        cursor.close()
        conn.close()
        
        if hasattr(st, "cache_data"):
            st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"Failed to insert test case into database: {e}")
        cursor.close()
        conn.close()
        return False

def get_crm_lock_from_db():
    conn = get_db_connection()
    if not conn:
        return "AVAILABLE", ""
    cursor = conn.cursor()
    cursor.execute('SELECT "crm_status", "locked_by" FROM "crm_machine_status" WHERE "lock_id" = 1')
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if row:
        return row[0], (row[1] if row[1] else "")
    return "AVAILABLE", ""

def update_crm_lock_in_db(status, user):
    conn = get_db_connection()
    if not conn:
        return
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE "crm_machine_status" 
        SET "crm_status" = %s, "locked_by" = %s, "last_updated" = CURRENT_TIMESTAMP
        WHERE "lock_id" = 1
    """, (status, user))
    conn.commit()
    cursor.close()
    conn.close()

def generate_professional_report_excel(df_data, report_title="UAT COMPLETED TEST CASES REPORT"):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Remove default sheet so we can add custom named tabs dynamically
    default_sheet = wb.active
    
    font_family = "Times New Roman"
    fill_dark_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_table_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_row_alt = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") # Light gray tint for zebra striping
    
    font_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=10, bold=True, color="000000")
    font_regular = Font(name=font_family, size=10, color="000000")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_side = Side(border_style="thin", color="BFBFBF")
    box_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    if df_data.empty:
        ws = wb.create_sheet(title="No Data")
        ws['A1'] = "No test cases available for the selected filters."
        ws['A1'].font = font_regular
        wb.remove(default_sheet)
        wb.save(output)
        output.seek(0)
        return output

    # Group data by Category, Module Name, and Path Type for granular tabs
    group_cols = [c for c in ['Category', 'Module Name', 'Path Type'] if c in df_data.columns]
    
    if len(group_cols) >= 3:
        grouped = df_data.groupby(['Category', 'Module Name', 'Path Type'])
    elif len(group_cols) >= 2:
        grouped = df_data.groupby(group_cols)
    else:
        grouped = [("All_Data", df_data)]

    used_sheet_names = set()

    for group_key, group_df in grouped:
        if isinstance(group_key, tuple) and len(group_key) == 3:
            cat, mod, path = group_key
        else:
            # Fallback if grouped differently: inspect the first row of the group dataframe
            cat = group_df.iloc[0].get('Category', 'General') if not group_df.empty else 'General'
            mod = group_df.iloc[0].get('Module Name', 'Module') if not group_df.empty else 'Module'
            path = group_df.iloc[0].get('Path Type', 'Positive') if not group_df.empty else 'Positive'

        # Safely determine if it is Positive or Negative based on path value
        path_str = str(path).lower().strip()
        is_neg = "neg" in path_str or "negative" in path_str
        path_short = "Neg" if is_neg else "Pos"
        
        base_name = f"{str(mod)[:20]} - {path_short}"
            
        # Clean invalid characters for Excel sheet names (\, /, ?, *, [, ], :)
        for char in ['\\', '/', '?', '*', '[', ']', ':']:
            base_name = base_name.replace(char, '')
            
        # Ensure sheet name is unique under Excel's 31-character limit without adding unwanted numbers if possible
        final_name = base_name[:31]
        counter = 1
        while final_name in used_sheet_names:
            # If it's a true duplicate name, ensure clean distinction
            suffix = f"_{counter}"
            final_name = base_name[:31 - len(suffix)] + suffix
            counter += 1
            
        used_sheet_names.add(final_name)
        
        ws = wb.create_sheet(title=final_name)
        ws.views.sheetView[0].showGridLines = True
        
        max_col = max(len(group_df.columns), 8)
        end_col_letter = get_column_letter(max_col)

        # Title Banner
        ws.merge_cells(f'A1:{end_col_letter}1')
        cell = ws['A1']
        cell.value = f"PEOPLE'S BANK — {report_title} ({final_name})"
        cell.font = font_title
        cell.fill = fill_dark_header
        cell.alignment = align_center
        ws.row_dimensions[1].height = 30

        ws['A2'] = f"Generated Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = font_regular
        ws.row_dimensions[2].height = 18

        ws.row_dimensions[4].height = 25
        headers = list(group_df.columns)
        for col_idx, h_text in enumerate(headers, 1):
            c = ws.cell(row=4, column=col_idx)
            c.value = h_text
            c.font = font_bold
            c.fill = fill_table_header
            c.alignment = align_center
            c.border = box_border

        fill_row_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_row_alt = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")

        curr_row = 5
        for idx, (_, r) in enumerate(group_df.iterrows()):
            ws.row_dimensions[curr_row].height = 24
            
            # Explicitly set white for odd rows, light gray for even rows
            row_fill = fill_row_alt if idx % 2 == 1 else fill_row_white

            for col_idx, col_name in enumerate(headers, 1):
                val = r.get(col_name, '')
                c = ws.cell(row=curr_row, column=col_idx, value=str(val) if val is not None else '')
                c.font = font_regular
                c.border = box_border
                c.fill = row_fill
                
                # Center-align Category, FE, SIBS, and other metadata columns
                if col_name in ['Category', 'TC ID', 'Path Type', 'Status', 'Executed Date', 'RRN', 'Utano', 'FE', 'SIBS', 'Severity', 'Priority', 'Defect Status']:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            curr_row += 1

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row >= 4:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 40), 12)

    # Remove the default blank sheet created by openpyxl
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    wb.save(output)
    output.seek(0)
    return output

def generate_official_defect_register_excel(defect_df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defect Tracking Register"
    
    ws.views.sheetView[0].showGridLines = True
    font_family = "Calibri"
    
    fill_main_banner = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_section_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_label_cell = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_table_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_row_alt = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
    
    font_banner_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=10, bold=True, color="1F4E78")
    font_bold = Font(name=font_family, size=10, bold=True, color="000000")
    font_regular = Font(name=font_family, size=10, color="000000")
    font_table_hdr = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_meta = Font(name=font_family, size=9, bold=True, color="595959")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_border_side = Side(border_style="thin", color="BFBFBF")
    box_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    ws['P1'] = "IT-TMP-027-Defect Tracking Register Template"
    ws['P1'].font = font_meta
    ws['P1'].alignment = align_right
    ws['P2'] = "Version: 1.0"
    ws['P2'].font = font_meta
    ws['P2'].alignment = align_right
    ws['P3'] = f"Effective Date: {datetime.today().strftime('%d/%m/%Y')}"
    ws['P3'].font = font_meta
    ws['P3'].alignment = align_right

    ws.merge_cells('E2:K2')
    cell = ws['E2']
    cell.value = "UAT DEFECT TRACKING REGISTER"
    cell.font = font_banner_title
    cell.fill = fill_main_banner
    cell.alignment = align_center
    ws.row_dimensions[2].height = 28

    ws['L4'] = "For the period from:"
    ws['L4'].font = font_bold
    ws['M4'] = datetime.today().strftime('%d/%m/%Y')
    ws['M4'].font = font_regular
    ws['O4'] = "to"
    ws['O4'].font = font_bold
    ws['P4'] = datetime.today().strftime('%d/%m/%Y')
    ws['P4'].font = font_regular

    ws.merge_cells('A6:S6')
    sec_cell = ws['A6']
    sec_cell.value = "   PART 1: PROJECT INFORMATION"
    sec_cell.font = font_section
    sec_cell.fill = fill_section_header
    sec_cell.alignment = align_left
    ws.row_dimensions[6].height = 22

    ws['A7'] = "Program Name"
    ws['A7'].font = font_bold
    ws['A7'].fill = fill_label_cell
    ws['A7'].border = box_border
    ws.merge_cells('B7:I7')
    ws['B7'] = "GRG CRM Banking Solution"
    ws['B7'].font = font_regular
    ws['B7'].border = box_border

    ws['J7'] = "Program #"
    ws['J7'].font = font_bold
    ws['J7'].fill = fill_label_cell
    ws['J7'].border = box_border
    ws.merge_cells('K7:L7')
    ws['K7'].border = box_border

    ws['M7'] = "IT PM"
    ws['M7'].font = font_bold
    ws['M7'].fill = fill_label_cell
    ws['M7'].border = box_border
    ws.merge_cells('N7:S7')
    ws['N7'].border = box_border

    ws['A8'] = "Project Name"
    ws['A8'].font = font_bold
    ws['A8'].fill = fill_label_cell
    ws['A8'].border = box_border
    ws.merge_cells('B8:I8')
    ws['B8'] = "People's Bank CRM Testing & Integration"
    ws['B8'].font = font_regular
    ws['B8'].border = box_border

    ws['J8'] = "Project #"
    ws['J8'].font = font_bold
    ws['J8'].fill = fill_label_cell
    ws['J8'].border = box_border
    ws.merge_cells('K8:L8')
    ws['K8'].border = box_border

    ws['M8'] = "Business PM"
    ws['M8'].font = font_bold
    ws['M8'].fill = fill_label_cell
    ws['M8'].border = box_border
    ws.merge_cells('N8:S8')
    ws['N8'].border = box_border

    ws['A9'] = "Project Sponsor"
    ws['A9'].font = font_bold
    ws['A9'].fill = fill_label_cell
    ws['A9'].border = box_border
    ws.merge_cells('B9:I9')
    ws['B9'] = "People's Bank IT Department"
    ws['B9'].font = font_regular
    ws['B9'].border = box_border

    ws['J9'] = "Project Manager"
    ws['J9'].font = font_bold
    ws['J9'].fill = fill_label_cell
    ws['J9'].border = box_border
    ws.merge_cells('K9:L9')
    ws['K9'].border = box_border

    ws['M9'] = "Project Start Date"
    ws['M9'].font = font_bold
    ws['M9'].fill = fill_label_cell
    ws['M9'].border = box_border
    ws.merge_cells('N9:S9')
    ws['N9'].border = box_border

    for r_idx in range(7, 10):
        ws.row_dimensions[r_idx].height = 20

    ws.merge_cells('A11:U11')
    def_sec = ws['A11']
    def_sec.value = "   PART 2: DEFECTS"
    def_sec.font = font_section
    def_sec.fill = fill_section_header
    def_sec.alignment = align_left
    ws.row_dimensions[11].height = 22

    headers = [
        "Origin (Build)", "Defect No.", "Defect Description", "Steps to Reproduce", 
        "Expected Results", "Defect Attachment", "CR Reference", "Application / Module", 
        "Defect Category", "Severity", "Priority", "Defect Status", "Detected By", 
        "Date of Defect Origin", "Assigned To", "Expected Date of Closure", 
        "Fixing Date", "Closed By", "Date of Closure", "SLA = (Date of Closure - Expected date of Closure)", "Comments"
    ]
    
    ws.row_dimensions[12].height = 32
    for col_idx, h_text in enumerate(headers, 1):
        c = ws.cell(row=12, column=col_idx)
        c.value = h_text
        c.font = font_table_hdr
        c.fill = fill_table_header
        c.alignment = align_center
        c.border = box_border

    curr_row = 13
    if not defect_df.empty:
        for idx, (_, r) in enumerate(defect_df.iterrows()):
            ws.row_dimensions[curr_row].height = 24
            
            # Fetch updated values directly from row with fallback to defaults
            desc_val = r.get("Defect Description") if str(r.get("Defect Description", "")).strip() else r.get("Test Case Description", "")
            steps_val = r.get("Test Steps", "") if str(r.get("Test Steps", "")).strip() else "Refer to system test case specification"
            exp_val = r.get("Expected Result", "") if str(r.get("Expected Result", "")).strip() else "System should process successfully without errors"

            row_data = [
                r.get("Origin (Build)", "CRM V1"),
                r.get("TC ID", ""),
                desc_val,
                steps_val,
                exp_val,
                safe_basename(r.get("Photo_Path", r.get("Receipt_Path", ""))),
                r.get("CR Reference", ""),
                r.get("Module Name", ""),
                r.get("Defect Category", r.get("Category", "")),
                r.get("Severity", ""),
                r.get("Priority", ""),
                r.get("Defect Status", ""),
                r.get("Detected By", r.get("Executed By", "")),
                r.get("Date of Defect Origin", r.get("Executed Date", "")),
                r.get("Assigned To", ""),
                r.get("Target Date", ""),
                r.get("Fixing Date", ""),
                r.get("Closed By", ""),
                r.get("Date of Closure", ""),
                r.get("SLA", ""),
                r.get("Comments", r.get("Remarks", ""))
            ]
            
            row_fill = fill_row_alt if idx % 2 == 1 else PatternFill(fill_type=None)

            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=curr_row, column=col_idx, value=str(val) if val is not None else '')
                c.font = font_regular
                c.border = box_border
                c.fill = row_fill
                c.alignment = align_left if col_idx in [3, 4, 5, 21] else align_center
            curr_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 12:
                continue
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(str(line)) > max_len:
                            max_len = len(str(line))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 45), 15)

    wb.save(output)
    output.seek(0)
    return output



















# def get_db_connection():
#     try:
#         db_url = st.secrets["postgres"]["url"]
#         conn = psycopg2.connect(db_url)
#         return conn
#     except Exception as e:
#         st.error(f"PostgreSQL Connection Failed: {e}")
#         return None

# # Page Configuration
# st.set_page_config(
#     page_title="People's Bank | GRG CRM UAT Portal",
#     page_icon="🏦",
#     layout="wide",
#     initial_sidebar_state="expanded"
# )

# UPLOAD_DIR = "uploads"
# os.makedirs(os.path.join(UPLOAD_DIR, "receipts"), exist_ok=True)
# os.makedirs(os.path.join(UPLOAD_DIR, "photos"), exist_ok=True)

# # ---------------------------------------------------------
# # ORACLE DATABASE CONNECTION & HELPER FUNCTIONS
# # ---------------------------------------------------------
# def get_db_connection():
#     try:
#         connection = oracledb.connect(
#             user="SYSTEM",
#             password="Dhinesh@98",
#             dsn="localhost:1521/FREEPDB1"
#         )
#         return connection
#     except Exception as e:
#         st.error(f"Database Connection Failed: {e}")
#         return None

# def clean_val(val):
#     if val is None:
#         return None
#     s = str(val).strip()
#     if not s or s.lower() in ["nan", "none", "", "nat"]:
#         return None
#     return s

# def load_data_from_db():
#     conn = get_db_connection()
#     if not conn:
#         return pd.DataFrame()
    
#     try:
#         query = """
#             SELECT 
#                 TC_ID as "TC ID", 
#                 CATEGORY as "Category",
#                 MODULE_NAME as "Module Name", 
#                 TEST_AREA as "Test Area", 
#                 TEST_CASE_DESCRIPTION as "Test Case Description", 
#                 PRE_CONDITIONS as "Pre-Conditions",
#                 TEST_STEPS as "Test Steps", 
#                 EXPECTED_RESULT as "Expected Result", 
#                 PATH_TYPE as "Path Type", 
#                 ACTUAL_RESULT as "Actual Result", 
#                 RRN as "RRN",
#                 UTANO as "Utano", 
#                 STATUS as "Status", 
#                 FE as "FE", 
#                 SIBS as "SIBS",
#                 REMARKS as "Remarks", 
#                 EXECUTED_BY as "Executed By", 
#                 EXECUTED_DATE as "Executed Date", 
#                 RECEIPT_PATH as "Receipt_Path",
#                 PHOTO_PATH as "Photo_Path",
#                 SEVERITY as "Severity",
#                 PRIORITY as "Priority",
#                 DEFECT_STATUS as "Defect Status",
#                 ASSIGNED_TO as "Assigned To",
#                 TARGET_DATE as "Target Date",
#                 ROOT_CAUSE as "Root Cause",
#                 DEFECT_DESCRIPTION as "Defect Description"
#             FROM uat_test_cases_v2
#         """
#         df = pd.read_sql(query, con=conn)
#         conn.close()
#         df = df.fillna("")
#         return df
#     except Exception as e:
#         if conn:
#             conn.close()
#         st.error(f"Error loading data from Oracle DB: {e}")
#         return pd.DataFrame()

# def derive_rrn(utano_val):
#     u_str = str(utano_val).strip()
#     if not u_str or u_str.lower() in ["nan", "none", "", "nat"]:
#         return "N/A"
#     if u_str.endswith(".0"):
#         u_str = u_str[:-2]
#     if len(u_str) > 6:
#         return u_str[6:]
#     return u_str


# def delete_test_case_from_db(tc_id, module_name):
#     conn = get_db_connection()
#     if not conn:
#         return False
#     cursor = conn.cursor()
#     try:
#         cursor.execute("""
#             DELETE FROM uat_test_cases_v2 
#             WHERE tc_id = :1 AND module_name = :2
#         """, (clean_val(tc_id), clean_val(module_name)))
#         conn.commit()
#         cursor.close()
#         conn.close()
#         return True
#     except Exception as e:
#         st.error(f"Failed to delete test case from Oracle DB: {e}")
#         cursor.close()
#         conn.close()
#         return False



# def safe_basename(path_val):
#     p_clean = clean_val(path_val)
#     if not p_clean:
#         return ""
#     try:
#         return os.path.basename(p_clean)
#     except Exception:
#         return ""

# def save_test_case_to_db(tc_id, module_name, status, actual_result, fe, sibs, utano, remarks, executed_by, executed_date, receipt_path, photo_path, defect_desc=""):
#     conn = get_db_connection()
#     if not conn:
#         return
#     cursor = conn.cursor()
    
#     c_actual = clean_val(actual_result)
#     c_fe = clean_val(fe)
#     c_sibs = clean_val(sibs)
#     c_utano = clean_val(utano)
    
#     rrn_val = derive_rrn(c_utano)
#     c_rrn = None if rrn_val == "N/A" else rrn_val
    
#     c_remarks = clean_val(remarks)
#     c_exec_by = clean_val(executed_by)
#     c_exec_date = clean_val(executed_date)
#     c_rec_path = clean_val(receipt_path)
#     c_pho_path = clean_val(photo_path)
#     c_def_desc = clean_val(defect_desc)
    
#     try:
#         if c_def_desc:
#             cursor.execute("""
#                 UPDATE uat_test_cases_v2 
#                 SET status = :1, actual_result = :2, fe = :3, sibs = :4, utano = :5, rrn = :6, 
#                     remarks = :7, executed_by = :8, executed_date = :9, receipt_path = :10, photo_path = :11,
#                     defect_description = :12, defect_status = 'Open'
#                 WHERE tc_id = :13 AND module_name = :14
#             """, (status, c_actual, c_fe, c_sibs, c_utano, c_rrn, c_remarks, c_exec_by, c_exec_date, c_rec_path, c_pho_path, c_def_desc, tc_id, module_name))
#         else:
#             cursor.execute("""
#                 UPDATE uat_test_cases_v2 
#                 SET status = :1, actual_result = :2, fe = :3, sibs = :4, utano = :5, rrn = :6, 
#                     remarks = :7, executed_by = :8, executed_date = :9, receipt_path = :10, photo_path = :11
#                 WHERE tc_id = :12 AND module_name = :13
#             """, (status, c_actual, c_fe, c_sibs, c_utano, c_rrn, c_remarks, c_exec_by, c_exec_date, c_rec_path, c_pho_path, tc_id, module_name))
#         conn.commit()
#     except Exception as e:
#         st.error(f"Database error updating test case: {e}")
#     cursor.close()
#     conn.close()

# def admin_update_full_defect_details(tc_id, module_name, test_steps, actual_result, executed_by, utano, fe, sibs, severity, priority, defect_status, assigned_to, target_date, root_cause, origin_build, defect_desc, defect_steps, defect_expected, defect_attachment, cr_ref, defect_cat, expected_date_closure, fixing_date, closed_by, date_closure, comments, date_defect_origin, detected_by):
#     conn = get_db_connection()
#     if not conn:
#         return
#     cursor = conn.cursor()
    
#     c_steps = clean_val(test_steps)
#     c_act = clean_val(actual_result)
#     c_exec_by = clean_val(executed_by)
#     c_utano = clean_val(utano)
#     rrn_val = derive_rrn(c_utano)
#     c_rrn = None if rrn_val == "N/A" else rrn_val
#     c_fe = clean_val(fe)
#     c_sibs = clean_val(sibs)
#     c_sev = clean_val(severity)
#     c_pri = clean_val(priority)
#     c_def_status = clean_val(defect_status)
#     c_assigned = clean_val(assigned_to)
#     c_target = clean_val(target_date)
#     c_root = clean_val(root_cause)

#     try:
#         cursor.execute("""
#             UPDATE uat_test_cases_v2 
#             SET test_steps = :1, actual_result = :2, executed_by = :3, utano = :4, rrn = :5, 
#                 fe = :6, sibs = :7, severity = :8, priority = :9, defect_status = :10, 
#                 assigned_to = :11, target_date = :12, root_cause = :13
#             WHERE tc_id = :14 AND module_name = :15
#         """, (c_steps, c_act, c_exec_by, c_utano, c_rrn, c_fe, c_sibs, c_sev, c_pri, c_def_status, c_assigned, c_target, c_root, tc_id, module_name))
#         conn.commit()
#     except Exception as e:
#         st.error(f"Database error updating defect details: {e}")
#     cursor.close()
#     conn.close()

# def insert_new_test_case_to_db(tc_id, category, module_name, test_area, test_desc, pre_cond, test_steps, exp_result, path_type):
#     conn = get_db_connection()
#     if not conn:
#         return False
#     cursor = conn.cursor()
#     try:
#         cursor.execute("""
#             INSERT INTO uat_test_cases_v2 (
#                 TC_ID, CATEGORY, MODULE_NAME, TEST_AREA, TEST_CASE_DESCRIPTION, 
#                 PRE_CONDITIONS, TEST_STEPS, EXPECTED_RESULT, PATH_TYPE, STATUS,
#                 SEVERITY, PRIORITY, DEFECT_STATUS, ASSIGNED_TO
#             ) VALUES (:1, :2, :3, :4, :5, :6, :7, :8, :9, 'PENDING', 'Medium', 'Medium', 'Open', 'Development Team')
#         """, (clean_val(tc_id), clean_val(category), clean_val(module_name), clean_val(test_area), 
#               clean_val(test_desc), clean_val(pre_cond), clean_val(test_steps), clean_val(exp_result), clean_val(path_type)))
#         conn.commit()
#         cursor.close()
#         conn.close()
        
#         # Clear cache/session state so the app immediately pulls the newly added test case from Oracle DB
#         if hasattr(st, "cache_data"):
#             st.cache_data.clear()
#         return True
#     except Exception as e:
#         st.error(f"Failed to insert test case into Oracle DB: {e}")
#         cursor.close()
#         conn.close()
#         return False

# def get_crm_lock_from_db():
#     conn = get_db_connection()
#     if not conn:
#         return "AVAILABLE", ""
#     cursor = conn.cursor()
#     cursor.execute("SELECT crm_status, locked_by FROM crm_machine_status WHERE lock_id = 1")
#     row = cursor.fetchone()
#     cursor.close()
#     conn.close()
#     if row:
#         return row[0], (row[1] if row[1] else "")
#     return "AVAILABLE", ""

# def update_crm_lock_in_db(status, user):
#     conn = get_db_connection()
#     if not conn:
#         return
#     cursor = conn.cursor()
#     cursor.execute("""
#         UPDATE crm_machine_status 
#         SET crm_status = :1, locked_by = :2, last_updated = CURRENT_TIMESTAMP
#         WHERE lock_id = 1
#     """, (status, user))
#     conn.commit()
#     cursor.close()
#     conn.close()

# def generate_professional_report_excel(df_data, report_title="UAT FILTERED REPORT"):
#     output = io.BytesIO()
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "UAT Report"
    
#     ws.views.sheetView[0].showGridLines = True
#     font_family = "Calibri"
    
#     fill_dark_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     fill_table_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
#     font_title = Font(name=font_family, size=13, bold=True, color="FFFFFF")
#     font_bold = Font(name=font_family, size=10, bold=True, color="000000")
#     font_regular = Font(name=font_family, size=10, color="000000")
    
#     align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
#     thin_side = Side(border_style="thin", color="BFBFBF")
#     box_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

#     max_col = max(len(df_data.columns), 8)
#     end_col_letter = get_column_letter(max_col)

#     ws.merge_cells(f'A1:{end_col_letter}1')
#     cell = ws['A1']
#     cell.value = f"PEOPLE'S BANK — {report_title}"
#     cell.font = font_title
#     cell.fill = fill_dark_header
#     cell.alignment = align_center
#     ws.row_dimensions[1].height = 30

#     ws['A2'] = f"Generated Date: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
#     ws['A2'].font = font_regular
#     ws.row_dimensions[2].height = 18

#     ws.row_dimensions[4].height = 25
#     headers = list(df_data.columns)
#     for col_idx, h_text in enumerate(headers, 1):
#         c = ws.cell(row=4, column=col_idx)
#         c.value = h_text
#         c.font = font_bold
#         c.fill = fill_table_header
#         c.alignment = align_center
#         c.border = box_border

#     curr_row = 5
#     for _, r in df_data.iterrows():
#         ws.row_dimensions[curr_row].height = 24
#         for col_idx, col_name in enumerate(headers, 1):
#             val = r.get(col_name, '')
#             c = ws.cell(row=curr_row, column=col_idx, value=str(val) if val is not None else '')
#             c.font = font_regular
#             c.border = box_border
#             if col_name in ['TC ID', 'Path Type', 'Status', 'Executed Date', 'RRN', 'Utano', 'Severity', 'Priority', 'Defect Status']:
#                 c.alignment = align_center
#             else:
#                 c.alignment = align_left
#         curr_row += 1

#     for col in ws.columns:
#         max_len = 0
#         col_letter = get_column_letter(col[0].column)
#         for cell in col:
#             if cell.row >= 4:
#                 val_str = str(cell.value or '')
#                 if len(val_str) > max_len:
#                     max_len = len(val_str)
#         ws.column_dimensions[col_letter].width = max(min(max_len + 4, 40), 12)

#     wb.save(output)
#     output.seek(0)
#     return output

# def generate_official_defect_register_excel(defect_df):
#     output = io.BytesIO()
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Defect Tracking Register"
    
#     ws.views.sheetView[0].showGridLines = True
#     font_family = "Calibri"
    
#     fill_dark_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     fill_yellow_banner = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
#     fill_gray_section = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
#     fill_table_header = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    
#     font_title = Font(name=font_family, size=11, bold=True, color="FFFFFF")
#     font_section = Font(name=font_family, size=10, bold=True, color="000000")
#     font_bold = Font(name=font_family, size=10, bold=True)
#     font_regular = Font(name=font_family, size=10)
    
#     align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
#     thin_border_side = Side(border_style="thin", color="000000")
#     box_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

#     ws.merge_cells('U1:U1')
#     cell = ws['A1']
#     cell.value = "UAT DEFECT TRACKING REGISTER"
#     cell.font = font_title
#     cell.fill = fill_dark_header
#     cell.alignment = align_center
#     ws.row_dimensions[1].height = 25

#     ws['R2'] = "IT-IMP-025: Defect Tracking Register Template"
#     ws['R2'].font = font_regular
#     ws['R3'] = "Version: 1.0"
#     ws['R3'].font = font_regular
#     ws['R4'] = f"Effective Date: {datetime.today().strftime('%d/%m/%Y')}"
#     ws['R4'].font = font_regular

#     ws.merge_cells('A5:U5')
#     cell = ws['A5']
#     cell.value = "PART I: PROJECT INFORMATION"
#     cell.font = font_section
#     cell.fill = fill_gray_section
#     cell.alignment = align_left
#     ws.row_dimensions[5].height = 20

#     ws['A6'] = "Project Name"
#     ws['A6'].font = font_bold
#     ws['A6'].fill = fill_yellow_banner
#     ws['A6'].border = box_border
    
#     ws.merge_cells('C6:E6')
#     ws['C6'] = "People's Bank CRM Testing & Integration"
#     ws['C6'].font = font_regular
#     ws['C6'].border = box_border

#     ws['P6'] = "Program #"
#     ws['P6'].font = font_bold
#     ws['P6'].fill = fill_yellow_banner
#     ws['P6'].border = box_border
#     ws['R6'] = "IT PM"
#     ws['R6'].font = font_regular
#     ws['R6'].border = box_border

#     ws.merge_cells('A7:B7')
#     ws['A7'] = "Project Sponsor"
#     ws['A7'].font = font_bold
#     ws['A7'].fill = fill_yellow_banner
#     ws['A7'].border = box_border
    
#     ws.merge_cells('C7:E7')
#     ws['C7'] = "People's Bank IT Department"
#     ws['C7'].font = font_regular
#     ws['C7'].border = box_border

#     ws['P7'] = "Project Manager"
#     ws['P7'].font = font_bold
#     ws['P7'].fill = fill_yellow_banner
#     ws['P7'].border = box_border
#     ws['R7'] = datetime.today().strftime('%d/%m/%Y')
#     ws['R7'].font = font_regular
#     ws['R7'].border = box_border

#     ws.merge_cells('A9:U9')
#     cell = ws['A9']
#     cell.value = "PART II: DEFECTS"
#     cell.font = font_section
#     cell.fill = fill_gray_section
#     cell.alignment = align_left
#     ws.row_dimensions[9].height = 20

#     headers = [
#         "Origin (Build)", "Defect No.", "Defect Description", "Steps to Reproduce", "Expected Results",
#         "Defect Attachment", "CR Reference", "Application / Module", "Defect Category", "Severity",
#         "Priority", "Defect Status", "Detected By", "Date of Defect Origin", "Assigned To",
#         "Expected Date of Closure", "Fixing Date", "Closed By", "Date of Closure",
#         "SLA = (Date of Closure - Expected date of Closure)", "Comments"
#     ]
    
#     ws.row_dimensions[10].height = 30
#     for col_idx, h_text in enumerate(headers, 1):
#         c = ws.cell(row=10, column=col_idx)
#         c.value = h_text
#         c.font = font_bold
#         c.fill = fill_table_header
#         c.alignment = align_center
#         c.border = box_border

#     curr_row = 11
#     for idx, r in defect_df.iterrows():
#         ws.row_dimensions[curr_row].height = 45
#         row_data = [
#             r.get('Origin (Build)', 'CRM V2'),
#             r.get('TC ID', ''),
#             r.get('Defect Description', r.get('Test Case Description', '')),
#             r.get('Steps to Reproduce', r.get('Test Steps', '')),
#             r.get('Expected Results', r.get('Expected Result', '')),
#             r.get('Defect Attachment', safe_basename(r.get('Photo_Path', r.get('Receipt_Path', '')))),
#             r.get('CR Reference', ''),
#             r.get('Module Name', ''),
#             r.get('Defect Category', r.get('Category', '')),
#             r.get('Severity', 'Medium'),
#             r.get('Priority', 'Medium'),
#             r.get('Defect Status', 'Open'),
#             r.get('Detected By', r.get('Executed By', '')),
#             r.get('Date of Defect Origin', r.get('Executed Date', '')),
#             r.get('Assigned To', 'Development Team'),
#             r.get('Expected Date of Closure', r.get('Target Date', '')),
#             r.get('Fixing Date', ''),
#             r.get('Closed By', ''),
#             r.get('Date of Closure', ''),
#             r.get('SLA', ''),
#             r.get('Comments', r.get('Remarks', ''))
#         ]
        
#         for col_idx, val in enumerate(row_data, 1):
#             c = ws.cell(row=curr_row, column=col_idx, value=val)
#             c.font = font_regular
#             c.border = box_border
#             if col_idx in [1, 2, 10, 11, 12, 14, 16, 17, 18, 19, 20]:
#                 c.alignment = align_center
#             else:
#                 c.alignment = align_left
#         curr_row += 1

#     for col in ws.columns:
#         max_len = 0
#         col_letter = get_column_letter(col[0].column)
#         for cell in col:
#             if cell.row > 4:
#                 val_str = str(cell.value or '')
#                 if len(val_str) > max_len:
#                     max_len = len(val_str)
#         ws.column_dimensions[col_letter].width = max(min(max_len + 3, 35), 12)

#     wb.save(output)
#     output.seek(0)
#     return output

# ---------------------------------------------------------
# STYLING (CSS)
# ---------------------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    
    .bank-header {
        background: linear-gradient(135deg, #061e3d 0%, #003366 50%, #0f4c81 100%);
        padding: 24px 30px;
        border-radius: 16px;
        color: white;
        margin-bottom: 25px;
        box-shadow: 0 10px 25px -5px rgba(0, 51, 102, 0.25);
        border-left: 6px solid #ffcc00;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .bank-header h1 { margin: 0; font-size: 24px; font-weight: 700; color: #ffffff; }
    .bank-header p { margin: 6px 0 0 0; color: #f1f5f9; font-size: 14px; opacity: 0.9; }
    .capacity-box {
        text-align: right; background: rgba(255,255,255,0.12);
        padding: 10px 18px; border-radius: 10px; min-width: 220px; white-space: nowrap;
    }
    .lock-card-available {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        border: 2px solid #22c55e; border-radius: 14px; padding: 16px 20px; margin-bottom: 20px;
    }
    .lock-card-inuse {
        background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%);
        border: 2px solid #ef4444; border-radius: 14px; padding: 16px 20px; margin-bottom: 20px;
    }
    .metric-card {
        background: #ffffff; border: 1px solid #e2e8f0; border-radius: 14px; padding: 20px; text-align: center;
        box-shadow: 0 4px 12px rgba(0,0,0,0.03);
    }
    .metric-num { font-size: 32px; font-weight: 800; line-height: 1.1; }
    .metric-label { font-size: 11px; font-weight: 700; text-transform: uppercase; color: #64748b; margin-top: 8px; }
    
    .execution-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.03);
    }
    .rrn-badge {
        background: #003366;
        color: #ffcc00;
        font-family: monospace;
        font-weight: 700;
        font-size: 14px;
        padding: 6px 12px;
        border-radius: 6px;
        display: inline-block;
        margin-top: 4px;
        letter-spacing: 0.5px;
    }
</style>
""", unsafe_allow_html=True)



st.markdown("""
    <style>
        /* Force Streamlit columns to stack vertically on mobile screens */
        @media (max-width: 768px) {
            [data-testid="stHorizontalBlock"] {
                flex-direction: column !important;
            }
            [data-testid="column"] {
                width: 100% !important;
                flex: 1 1 100% !important;
                min-width: auto !important;
                margin-bottom: 0.5rem !important;
            }
            h1 {
                font-size: 1.4rem !important;
            }
            p, span {
                font-size: 0.95rem !important;
            }
        }

        /* Make the main page container a flex container to lock the footer at the bottom */
        .stApp {
            display: flex;
            flex-direction: column;
            min-height: 100vh;
        }
        
        .main {
            flex: 1;
        }

        /* Static, non-floating footer that rests naturally below content */
        .custom-footer {
            width: 100%;
            text-align: center;
            padding: 20px 0;
            font-family: 'Times New Roman', Times, serif;
            font-size: 0.9rem;
            color: #555555;
            border-top: 1px solid #EAEAEA;
            margin-top: auto;
            background-color: inherit;
        }
        
        .stDataFrame { width: 100%; overflow-x: auto; }
        .stButton button, .stDownloadButton button { width: 100%; }
    </style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# SESSION STATE & DB DATA
# ---------------------------------------------------------
if 'authenticated_role' not in st.session_state:
    st.session_state.authenticated_role = None
if 'logged_user' not in st.session_state:
    st.session_state.logged_user = ""

db_crm_status, db_crm_user = get_crm_lock_from_db()
df = load_data_from_db()

# ---------------------------------------------------------
# SIDEBAR — ROLES & AUTHENTICATION
# ---------------------------------------------------------
try:
    st.sidebar.image("peoples_bank_logo.jpeg", width=180)
except Exception:
    st.sidebar.markdown("<h3 style='color:#003366; margin:0;'>PEOPLE'S BANK</h3>", unsafe_allow_html=True)

st.sidebar.caption("Sri Lanka's Leader in Banking & Technology")
st.sidebar.divider()

USERS_DB = {
    "admin": {"password": "admin123", "role": "Admin / Manager"},
    "tester": {"password": "tester123", "role": "Tester"},
    "developer": {"password": "dev123", "role": "Developer"},
    "qa": {"password": "qa123", "role": "QA Engineer"}
}

st.sidebar.markdown("### 🔑 System Login")

if st.session_state.authenticated_role is None:
    username = st.sidebar.text_input("Username").strip().lower()
    password = st.sidebar.text_input("Password", type="password")
    
    if st.sidebar.button("Login", use_container_width=True):
        if username in USERS_DB and USERS_DB[username]["password"] == password:
            st.session_state.authenticated_role = USERS_DB[username]["role"]
            st.session_state.logged_user = username.upper()
            st.rerun()
        else:
            st.sidebar.error("Invalid Username or Password")
    
    # st.sidebar.info("""
    # **Roles & Demo Credentials:**
    # * **Admin/Manager:** `admin` / `admin123`
    # * **Tester:** `tester` / `tester123`
    # * **Developer:** `developer` / `dev123`
    # * **QA Engineer:** `qa` / `qa123`
    # """)
    menu = "📊 Live Dashboard"
else:
    st.sidebar.success(f"👤 **{st.session_state.logged_user}**\n\nRole: `{st.session_state.authenticated_role}`")
    if st.sidebar.button("Logout", use_container_width=True):
        st.session_state.authenticated_role = None
        st.session_state.logged_user = ""
        st.rerun()

    current_role = st.session_state.authenticated_role
    if current_role == "Admin / Manager":
        menu = st.sidebar.radio("Navigation", ["📊 Live Dashboard", "🧪 Test Execution & Scenarios", "🛠️ Defect Tracker", "⚙️ Admin Management", "📄 Reports"])
    elif current_role == "Tester":
        menu = st.sidebar.radio("Navigation", ["📊 Live Dashboard", "🧪 Test Execution & Scenarios", "🛠️ Defect Tracker", "📄 Reports"])
    elif current_role == "Developer":
        menu = st.sidebar.radio("Navigation", ["🛠️ Defect Tracker", "📊 Live Dashboard", "📄 Reports"])
    else:
        menu = st.sidebar.radio("Navigation", ["📊 Live Dashboard", "🛠️ Defect Tracker", "📄 Reports"])

# Header Banner
st.markdown("""
<div class="bank-header">
    <div>
        <h1>PEOPLE'S BANK — GRG CRM UAT PORTAL</h1>
        <p>Cash Recycling Machine (CRM) Final User Acceptance Testing Portal (Oracle DB Enabled)</p>
    </div>
    <div class="capacity-box">
        <span style="font-size: 11px; font-weight: 600; color: #ffcc00; text-transform: uppercase;">Team Capacity</span><br>
        <span style="font-size: 15px; font-weight: 700; color: #ffffff;">12 Members (1 CRM)</span>
    </div>
</div>
""", unsafe_allow_html=True)

# CRM Lock Widget
col_lock, col_act = st.columns([3, 1])
with col_lock:
    if db_crm_status == "AVAILABLE":
        st.markdown("""
        <div class="lock-card-available">
            <div style="font-size: 13px; font-weight: 700; color: #15803d;">🟢 CRM MACHINE STATUS</div>
            <p style="font-size: 20px; font-weight: 800; color: #166534; margin: 0;">READY & AVAILABLE FOR TESTING</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="lock-card-inuse">
            <div style="font-size: 13px; font-weight: 700; color: #b91c1c;">🔴 CRM MACHINE STATUS — LOCKED</div>
            <p style="font-size: 20px; font-weight: 800; color: #991b1b; margin: 0;">IN USE BY: {db_crm_user.upper()}</p>
        </div>
        """, unsafe_allow_html=True)

with col_act:
    st.write("")
    with st.popover("⚙️ Lock / Release CRM", use_container_width=True):
        tester_input = st.text_input("Tester / Pair Name", value=db_crm_user)
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🔒 Lock"):
                if tester_input:
                    update_crm_lock_in_db("IN_USE", tester_input)
                    st.rerun()
        with c2:
            if st.button("🔓 Release"):
                update_crm_lock_in_db("AVAILABLE", "")
                st.rerun()

st.divider()

# ---------------------------------------------------------
# 1. LIVE DASHBOARD
# ---------------------------------------------------------
if menu == "📊 Live Dashboard":
    st.subheader("🎯 Overall Testing Progress Summary")
    
    # Normalize status values safely
    dash_df = df.copy()
    if not dash_df.empty and 'Status' in dash_df.columns:
        dash_df['Normalized_Status'] = dash_df['Status'].apply(
            lambda s: str(s).upper().strip() if str(s).strip() and str(s).lower() not in ['nan', 'none', ''] else 'PENDING'
        )
    else:
        dash_df['Normalized_Status'] = 'PENDING'

    total = len(dash_df)
    passed = len(dash_df[dash_df['Normalized_Status'] == 'PASS'])
    failed = len(dash_df[dash_df['Normalized_Status'] == 'FAIL'])
    NA = len(dash_df[dash_df['Normalized_Status'] == 'N/A'])
    pending = len(dash_df[dash_df['Normalized_Status'] == 'PENDING'])
    pass_pct = round((passed / total * 100), 1) if total > 0 else 0

    # Global Metrics Cards
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.markdown(f'<div class="metric-card"><div class="metric-num">{total}</div><div class="metric-label">Total Cases</div></div>', unsafe_allow_html=True)
    m2.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #16a34a;">{passed}</div><div class="metric-label">Passed ({pass_pct}%)</div></div>', unsafe_allow_html=True)
    m3.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #dc2626;">{failed}</div><div class="metric-label">Failed</div></div>', unsafe_allow_html=True)
    m4.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #d97706;">{NA}</div><div class="metric-label">Not Applicable</div></div>', unsafe_allow_html=True)
    m5.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #64748b;">{pending}</div><div class="metric-label">Pending</div></div>', unsafe_allow_html=True)

    st.write("")
    st.divider()

    # --- CATEGORY & MODULE BREAKDOWN SECTION ---
    st.markdown("### 🏷️ Category & Module Execution Breakdown")
    
    categories = list(dash_df['Category'].dropna().unique()) if 'Category' in dash_df.columns else []
    
    if categories:
        cat_tabs = st.tabs([f"📂 {cat}" for cat in categories])
        
        for idx, cat in enumerate(categories):
            with cat_tabs[idx]:
                cat_subset = dash_df[dash_df['Category'] == cat]
                
                c_total = len(cat_subset)
                c_pass = len(cat_subset[cat_subset['Normalized_Status'] == 'PASS'])
                c_fail = len(cat_subset[cat_subset['Normalized_Status'] == 'FAIL'])
                c_pend = len(cat_subset[cat_subset['Normalized_Status'] == 'PENDING'])
                c_na = len(cat_subset[cat_subset['Normalized_Status'] == 'N/A'])
                c_rate = (c_pass / c_total * 100) if c_total > 0 else 0.0
                
                sub_col1, sub_col2, sub_col3, sub_col4, sub_col5 = st.columns(5)
                sub_col1.markdown(f'<div class="metric-card"><div class="metric-num">{c_total}</div><div class="metric-label">Total Cases</div></div>', unsafe_allow_html=True)
                sub_col2.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #16a34a;">{c_pass}</div><div class="metric-label">Passed ({c_rate:.1f}%)</div></div>', unsafe_allow_html=True)
                sub_col3.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #dc2626;">{c_fail}</div><div class="metric-label">Failed</div></div>', unsafe_allow_html=True)
                sub_col4.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #d97706;">{c_na}</div><div class="metric-label">Not Applicable</div></div>', unsafe_allow_html=True)
                sub_col5.markdown(f'<div class="metric-card"><div class="metric-num" style="color: #64748b;">{c_pend}</div><div class="metric-label">Pending</div></div>', unsafe_allow_html=True)
                
                st.write("")
                st.markdown("##### Module Performance Summary Table")
                
                if 'Module Name' in cat_subset.columns:
                    module_summary = cat_subset.groupby('Module Name')['Normalized_Status'].value_counts().unstack(fill_value=0)
                    
                    for st_col in ['PASS', 'FAIL', 'PENDING', 'N/A']:
                        if st_col not in module_summary.columns:
                            module_summary[st_col] = 0
                            
                    module_summary['Total'] = module_summary[['PASS', 'FAIL', 'PENDING', 'N/A']].sum(axis=1)
                    module_summary['Pass Rate (%)'] = (module_summary['PASS'] / module_summary['Total'] * 100).round(1)
                    
                    display_summary = module_summary[['Total', 'PASS', 'FAIL', 'PENDING', 'N/A', 'Pass Rate (%)']].reset_index()
                    
                    st.dataframe(
                        display_summary,
                        use_container_width=True,
                        hide_index=True
                    )

    st.divider()
    st.subheader("📋 Master Test Cases Overview & Multi-Filters")
    
    all_cats_list = list(df['Category'].unique()) if not df.empty and 'Category' in df.columns else []
    all_mods_list = list(df['Module Name'].unique()) if not df.empty and 'Module Name' in df.columns else []
    
    f1, f2, f3, f4, f5 = st.columns(5)
    with f1:
        sel_cats = st.multiselect("Categories", all_cats_list, default=all_cats_list, key="dash_cats")
    with f2:
        sel_mods = st.multiselect("Modules / Sheets", all_mods_list, default=all_mods_list, key="dash_mods")
    with f3:
        sel_paths = st.multiselect("Path Types", ["Positive", "Negative"], default=["Positive", "Negative"], key="dash_paths")
    with f4:
        sel_statuses = st.multiselect("Statuses", ["PASS", "FAIL", "N/A", "PENDING"], default=["PASS", "FAIL", "N/A", "PENDING"], key="dash_statuses")
    with f5:
        search_kw = st.text_input("Search ID / Desc / Utano / RRN", placeholder="e.g. CD-01 or Utano/RRN")

    d_col1, d_col2 = st.columns(2)
    with d_col1:
        enable_date_filter = st.checkbox("Enable Execution Date Filter", key="dash_date_chk")
    with d_col2:
        if enable_date_filter:
            date_range = st.date_input("Select Date Range", value=(date.today(), date.today()), key="dash_date_rng")
        else:
            date_range = None

    filtered_df = df.copy()
    if not filtered_df.empty:
        if 'Category' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Category'].isin(sel_cats)]
        if 'Module Name' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Module Name'].isin(sel_mods)]
        if 'Path Type' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Path Type'].isin(sel_paths)]
        if 'Status' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Status'].isin(sel_statuses)]
        if search_kw:
            filtered_df = filtered_df[
                filtered_df['TC ID'].str.contains(search_kw, case=False, na=False) | 
                filtered_df['Test Case Description'].str.contains(search_kw, case=False, na=False) | 
                filtered_df['Utano'].str.contains(search_kw, case=False, na=False) | 
                filtered_df['RRN'].str.contains(search_kw, case=False, na=False)
            ]
        
        if enable_date_filter and isinstance(date_range, tuple) and len(date_range) == 2:
            start_d, end_d = date_range
            def parse_dt(val):
                if not val or pd.isna(val): return None
                try:
                    return pd.to_datetime(val).date()
                except Exception:
                    return None
            
            exec_dates = filtered_df['Executed Date'].apply(parse_dt)
            mask = exec_dates.apply(lambda d: d is not None and start_d <= d <= end_d)
            filtered_df = filtered_df[mask]

    display_columns = [
        'TC ID', 'Test Area', 'Test Case Description', 'Pre-Conditions', 
        'Test Steps', 'Expected Result', 'Actual Result', 'RRN', 'Utano', 'Status', 'FE', 'SIBS', 'Executed Date', 'Remarks'
    ]
    st.dataframe(filtered_df[display_columns] if not filtered_df.empty else filtered_df, use_container_width=True, height=450, hide_index=True)

    dash_export_df = filtered_df[display_columns] if not filtered_df.empty else pd.DataFrame(columns=display_columns)
    dash_buf = generate_professional_report_excel(dash_export_df, report_title="CRM DASHBOARD EXPORT REPORT")
    
    st.download_button(
        label="📥 Download Filtered Dashboard Report (.xlsx)",
        data=dash_buf.getvalue(),
        file_name=f"PeoplesBank_CRM_Dashboard_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# ---------------------------------------------------------
# 2. TEST EXECUTION & SCENARIOS
# ---------------------------------------------------------
elif menu == "🧪 Test Execution & Scenarios":
    st.subheader("🧪 Test Execution Panel — Positive & Negative Scenarios")
    
    can_execute = st.session_state.authenticated_role in ["Admin / Manager", "Tester"]
    if not can_execute:
        st.warning("⚠️ You are logged in as a Viewer/Developer role. Test execution and status updates are restricted to **Testers** and **Admin / Managers**.")
    
    categories = list(df['Category'].unique()) if not df.empty and 'Category' in df.columns else []
    if not categories:
        st.warning("No test cases found in DB. Please run your import script first.")
    else:
        col_sel1, col_sel2, col_sel3 = st.columns(3)
        with col_sel1:
            f_cat = st.selectbox("Select Category", categories, key="exec_cat_sel")
            cat_df = df[df['Category'] == f_cat] if not df.empty else pd.DataFrame()
        with col_sel2:
            modules = list(cat_df['Module Name'].unique()) if not cat_df.empty and 'Module Name' in cat_df.columns else []
            f_mod = st.selectbox("Select Module / Sheet", modules, key="exec_mod_sel")
            mod_df = cat_df[cat_df['Module Name'] == f_mod] if not cat_df.empty else pd.DataFrame()
        with col_sel3:
            exec_status_filter = st.selectbox("Filter by Status", ["All", "PENDING", "PASS", "FAIL", "N/A"], key="exec_stat_filt")
        
        if exec_status_filter != "All" and not mod_df.empty:
            mod_df['Normalized_Status'] = mod_df['Status'].apply(
                lambda s: str(s).upper().strip() if str(s).strip() and str(s).lower() not in ['nan', 'none', ''] else 'PENDING'
            )
            mod_df = mod_df[mod_df['Normalized_Status'] == exec_status_filter]

        exec_search_kw = st.text_input("🔍 Search within Module (by ID, Desc, Tester, Utano, or RRN)", placeholder="e.g. CD-01, John, Utano, etc.", key="exec_search_input")
        if exec_search_kw and not mod_df.empty:
            mod_df = mod_df[
                mod_df['TC ID'].str.contains(exec_search_kw, case=False, na=False) | 
                mod_df['Test Case Description'].str.contains(exec_search_kw, case=False, na=False) | 
                mod_df['Executed By'].str.contains(exec_search_kw, case=False, na=False) | 
                mod_df['Utano'].str.contains(exec_search_kw, case=False, na=False) | 
                mod_df['RRN'].str.contains(exec_search_kw, case=False, na=False)
            ]
        
        if not mod_df.empty and 'Path Type' in mod_df.columns:
            path_type_str = mod_df.iloc[0]['Path Type']
            st.markdown(f"### **{f_cat} ➔ {f_mod} ({path_type_str} Scenarios) [Showing: {exec_status_filter}]**")
        else:
            st.info("No test cases match the selected status filter or search query in this module.")

        for idx, row in mod_df.iterrows():
            status_val = str(row.get('Status', 'PENDING')).upper().strip()
            badge = "🟢" if status_val == 'PASS' else ("🔴" if status_val == 'FAIL' else ("🟡" if status_val == 'N/A' else "🔵"))
            
            with st.expander(f"{badge} [{row.get('Status', 'PENDING')}] {row['TC ID']} — {row['Test Case Description']}"):
                st.markdown(f"**Test Area:** `{row['Test Area']}`")
                
                col_info1, col_info2 = st.columns(2)
                with col_info1:
                    pre_cond_text = row['Pre-Conditions'] if row.get('Pre-Conditions') else "None"
                    st.markdown("**Pre-Conditions:**")
                    st.info(pre_cond_text)
                    
                    steps_text = row['Test Steps'] if row.get('Test Steps') else "No steps specified"
                    st.markdown("**Test Steps:**")
                    st.warning(steps_text)
                with col_info2:
                    exp_text = row['Expected Result'] if row.get('Expected Result') else "No expected result specified"
                    st.markdown("**Expected Result:**")
                    st.success(exp_text)
                
                st.divider()

                unique_suffix = f"{row['Module Name']}_{row['TC ID']}_{idx}".replace(" ", "_")

                st.markdown('<div class="execution-card">', unsafe_allow_html=True)
                
                # Wrap execution inputs inside a native st.form to completely eliminate unwanted reruns on file uploads
                with st.form(key=f"exec_form_{unique_suffix}"):
                    col_left, col_right = st.columns(2, gap="large")
                    
                    with col_left:
                        st.markdown("#### 📝 Status & Transaction Details")
                        
                        status_opt = ["PENDING", "PASS", "FAIL", "N/A"]
                        st_curr = row.get('Status', 'PENDING')
                        if st_curr not in status_opt:
                            st_curr = "PENDING"
                        new_status = st.selectbox("Execution Status", status_opt, index=status_opt.index(st_curr), key=f"st_{unique_suffix}", disabled=not can_execute)
                        
                        exec_val = "" if str(row.get('Executed By')) in ["nan", "None", ""] else row['Executed By']
                        tester_name = st.text_input("Executed By (Tester)", value=exec_val, key=f"ex_{unique_suffix}", disabled=not can_execute)
                        
                        existing_date_str = str(row.get('Executed Date', ''))
                        try:
                            default_exec_date = datetime.strptime(existing_date_str.split()[0], "%Y-%m-%d").date() if existing_date_str else date.today()
                        except Exception:
                            default_exec_date = date.today()
                        
                        selected_exec_date = st.date_input("Execution Date", value=default_exec_date, key=f"date_{unique_suffix}", disabled=not can_execute)
                        
                        uano_val = "" if str(row.get('Utano')) in ["nan", "None", ""] else row['Utano']
                        new_utano = st.text_input("UTANO (from bill)", value=uano_val, key=f"utano_{unique_suffix}", disabled=not can_execute)
                        
                        calc_rrn = derive_rrn(new_utano)
                        st.markdown(f"""
                            <div style="background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 8px; padding: 10px 14px; margin-bottom: 15px;">
                                <span style="font-size: 11px; font-weight: 700; color: #475569; text-transform: uppercase;">Auto-Calculated RRN:</span><br>
                                <div class="rrn-badge">{calc_rrn}</div>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        fe_val = "" if str(row.get('FE')) in ["nan", "None", ""] else row['FE']
                        new_fe = st.text_input("FE", value=fe_val, key=f"fe_{unique_suffix}", disabled=not can_execute)
                        
                        sibs_val = "" if str(row.get('SIBS')) in ["nan", "None", ""] else row['SIBS']
                        new_sibs = st.text_input("SIBS", value=sibs_val, key=f"sibs_{unique_suffix}", disabled=not can_execute)

                    with col_right:
                        st.markdown("#### 📋 Results & Attachments")
                        
                        act_val = "" if str(row.get('Actual Result')) in ["nan", "None", ""] else row['Actual Result']
                        act_res = st.text_area("Actual Result", value=act_val, height=105, key=f"act_{unique_suffix}", disabled=not can_execute)
                        
                        rem_val = "" if str(row.get('Remarks')) in ["nan", "None", ""] else row['Remarks']
                        remarks = st.text_input("Remarks", value=rem_val, key=f"rem_{unique_suffix}", disabled=not can_execute)
                        
                        st.write("")
                        rc_col, ph_col = st.columns(2)
                        
                        # --- RECEIPT UPLOAD & PREVIEW ---
                        with rc_col:
                            st.markdown("**Receipt (.jpg/.pdf)**")
                            receipt_file = st.file_uploader(
                                "Upload Receipt", 
                                type=["jpg", "png", "jpeg", "pdf"], 
                                key=f"rec_{unique_suffix}", 
                                disabled=not can_execute,
                                label_visibility="collapsed"
                            )
                            
                            existing_receipt = row.get('Receipt_Path', '')
                            if existing_receipt and os.path.exists(str(existing_receipt)):
                                st.caption(f"📎 Current: `{safe_basename(existing_receipt)}`")
                                if str(existing_receipt).lower().endswith(('.jpg', '.jpeg', '.png')):
                                    st.image(existing_receipt, width=150, caption="Receipt Preview")
                                
                                if can_execute and st.checkbox("🗑️ Remove Receipt", key=f"rm_rec_{unique_suffix}"):
                                    existing_receipt = "REMOVE" # Flag for removal

                        # --- ERROR PHOTO UPLOAD & PREVIEW ---
                        with ph_col:
                            st.markdown("**Error Photo (.jpg)**")
                            photo_file = st.file_uploader(
                                "Upload Error Photo", 
                                type=["jpg", "png", "jpeg"], 
                                key=f"pho_{unique_suffix}", 
                                disabled=not can_execute,
                                label_visibility="collapsed"
                            )
                            
                            existing_photo = row.get('Photo_Path', '')
                            if existing_photo and os.path.exists(str(existing_photo)):
                                st.caption(f"📷 Current: `{safe_basename(existing_photo)}`")
                                st.image(existing_photo, width=150, caption="Error Photo Preview")
                                
                                if can_execute and st.checkbox("🗑️ Remove Photo", key=f"rm_pho_{unique_suffix}"):
                                    existing_photo = "REMOVE" # Flag for removal

                    st.divider()
                    st.markdown("#### 🐞 Defect Description")
                    def_desc_val = row.get('Defect Description', '')
                    new_def_desc = st.text_area(
                        "Enter Defect Description (if any defect occurred during testing):", 
                        value=def_desc_val, 
                        height=80, 
                        key=f"def_desc_{unique_suffix}", 
                        disabled=not can_execute
                    )

                    st.write("")
                    submitted = st.form_submit_button(f"💾 Save & Sync to Database ({row['TC ID']})", use_container_width=True, disabled=not can_execute)
                    
                    if submitted:
                        # Handle Receipt Path saving/removal
                        rec_path = row.get('Receipt_Path', '')
                        if locals().get('existing_receipt') == "REMOVE":
                            rec_path = ""
                        if receipt_file is not None:
                            rec_path = os.path.join(UPLOAD_DIR, "receipts", f"{row['TC ID']}_{receipt_file.name}")
                            with open(rec_path, "wb") as f:
                                f.write(receipt_file.getbuffer())

                        # Handle Photo Path saving/removal
                        pho_path = row.get('Photo_Path', '')
                        if locals().get('existing_photo') == "REMOVE":
                            pho_path = ""
                        if photo_file is not None:
                            pho_path = os.path.join(UPLOAD_DIR, "photos", f"{row['TC ID']}_{photo_file.name}")
                            with open(pho_path, "wb") as f:
                                f.write(photo_file.getbuffer())

                        exec_date_str = selected_exec_date.strftime("%Y-%m-%d %H:%M:%S")

                        save_test_case_to_db(
                            tc_id=row['TC ID'], 
                            module_name=row['Module Name'], 
                            status=new_status, 
                            actual_result=act_res, 
                            fe=new_fe, 
                            sibs=new_sibs, 
                            utano=new_utano, 
                            remarks=remarks, 
                            executed_by=tester_name, 
                            executed_date=exec_date_str, 
                            receipt_path=rec_path, 
                            photo_path=pho_path,
                            defect_desc=new_def_desc
                        )
                        st.success(f"Successfully updated test case {row['TC ID']} in database!")
                        st.rerun()
                        
                st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------------------------------------
# 3. DEFECT TRACKER
# ---------------------------------------------------------
elif menu == "🛠️ Defect Tracker":
    st.subheader("🛠️ Centralized Defect Tracker Report")
    
    if not df.empty and 'Status' in df.columns:
        # Only include FAIL (exclude BLOCKED and N/A)
        defect_df = df[df['Status'] == 'FAIL'].copy()
    else:
        defect_df = pd.DataFrame()

    # --- SORT & SEARCH DEFECTS ---
    if not defect_df.empty and 'TC ID' in defect_df.columns:
        defect_df = defect_df.sort_values(by='TC ID', ascending=True).reset_index(drop=True)

    if not defect_df.empty:
        st.markdown("#### 🔍 Search & Filter Defects")
        sch_col1, sch_col2 = st.columns([2, 2])
        with sch_col1:
            def_search_kw = st.text_input("Search by Defect ID (TC ID) or Description", placeholder="e.g. TC_CD_CB or Invalid", key="def_search_input")
        with sch_col2:
            enable_def_date_filter = st.checkbox("Enable Execution Date Filter for Defects", key="def_date_chk")

        # Apply search keyword filter
        if def_search_kw and not defect_df.empty:
            defect_df = defect_df[
                defect_df['TC ID'].str.contains(def_search_kw, case=False, na=False) | 
                defect_df['Test Case Description'].str.contains(def_search_kw, case=False, na=False) |
                defect_df.get('Defect Description', pd.Series('', index=defect_df.index)).str.contains(def_search_kw, case=False, na=False)
            ]

        # --- DATE FILTER CONFIGURATION FOR DEFECTS ---
        if enable_def_date_filter:
            def_date_mode = st.radio("Date Filter Mode", ["Date Range", "Specific Date"], horizontal=True, key="def_date_mode")
            def_date_range = None
            specific_def_date = None

            if def_date_mode == "Date Range":
                def_date_range = st.date_input("Select Execution Date Range", value=(date.today(), date.today()), key="def_date_rng")
            else:
                specific_def_date = st.date_input("Select Specific Execution Date", value=date.today(), key="def_specific_date")

            def parse_dt(val):
                if not val or pd.isna(val): return None
                try:
                    return pd.to_datetime(val).date()
                except Exception:
                    return None
            
            exec_dates = defect_df['Executed Date'].apply(parse_dt)
            
            if def_date_mode == "Date Range" and isinstance(def_date_range, tuple) and len(def_date_range) == 2:
                start_d, end_d = def_date_range
                mask = exec_dates.apply(lambda d: d is not None and start_d <= d <= end_d)
                defect_df = defect_df[mask]
            elif def_date_mode == "Specific Date" and specific_def_date:
                mask = exec_dates.apply(lambda d: d is not None and d == specific_def_date)
                defect_df = defect_df[mask]

    if defect_df.empty:
        st.success("🎉 No failed test cases match your criteria.")
    else:
        st.warning(f"⚠️ Total Active Defects / Failed Tests Matching Criteria: {len(defect_df)}")
        
        is_admin = (st.session_state.authenticated_role == "Admin / Manager")
        
        if is_admin:
            st.markdown("### 📥 Download Official UAT Defect Tracking Register")
            
            excel_buffer = generate_official_defect_register_excel(defect_df)
            
            st.download_button(
                label="📥 Download Official UAT Defect Tracking Register (.xlsx)",
                data=excel_buffer.getvalue(),
                file_name=f"PeoplesBank_UAT_Defect_Tracking_Register_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.divider()

        st.markdown("### 📋 Active Defects Overview Table")
        
        tracker_display_cols = [
            'TC ID', 'Module Name', 'Path Type', 'Status', 'Severity', 
            'Priority', 'Defect Status', 'Assigned To', 'Defect Description', 
            'Executed By', 'Executed Date', 'Utano', 'RRN'
        ]
        
        for col in tracker_display_cols:
            if col not in defect_df.columns:
                defect_df[col] = ""

        st.dataframe(
            defect_df[tracker_display_cols], 
            use_container_width=True, 
            height=350, 
            hide_index=True
        )
        
        st.divider()
        st.markdown("### ⚙️ Individual Defect Inspection & Admin Management")

        for idx, row in defect_df.iterrows():
            tc_id = row['TC ID']
            mod_name = row['Module Name']
            
            # Fallbacks to ensure descriptions and steps are never empty
            default_desc = row.get('Defect Description') if str(row.get('Defect Description', '')).strip() else row.get('Test Case Description', '')
            default_steps = row.get('Steps to Reproduce') if str(row.get('Steps to Reproduce', '')).strip() else row.get('Test Steps', '')
            default_expected = row.get('Expected Results') if str(row.get('Expected Results', '')).strip() else row.get('Expected Result', '')
            
            with st.expander(f"🔴 [{tc_id}] {default_desc} — Status: {row.get('Status')} (Module: {mod_name})"):
                
                d_key = f"def_{mod_name}_{tc_id}_{idx}"
                
                if is_admin:
                    st.markdown("#### ⚙️ Admin Complete Defect Editing & Matrix Panel")
                    
                    # Wrap admin edit panel inside an st.form so input values are properly captured and saved on button click
                    with st.form(key=f"admin_defect_form_{d_key}"):
                        col_a1, col_a2, col_a3 = st.columns(3)
                        with col_a1:
                            adm_origin_build = st.text_input("Origin (Build)", value=row.get('Origin (Build)', 'CRM V1'), key=f"orig_build_{d_key}")
                            adm_cr_ref = st.text_input("CR Reference", value=row.get('CR Reference', ''), key=f"cr_ref_{d_key}")
                            adm_defect_cat = st.text_input("Defect Category", value=row.get('Defect Category', row.get('Category', '')), key=f"def_cat_{d_key}")
                        with col_a2:
                            adm_fixing_date = st.text_input("Fixing Date", value=row.get('Fixing Date', ''), key=f"fix_date_{d_key}")
                            adm_closed_by = st.text_input("Closed By", value=row.get('Closed By', ''), key=f"closed_by_{d_key}")
                            adm_date_closure = st.text_input("Date of Closure", value=row.get('Date of Closure', ''), key=f"date_closure_{d_key}")
                        with col_a3:
                            adm_sla = st.text_input("SLA", value=row.get('SLA', ''), key=f"sla_{d_key}")
                            adm_detected_by = st.text_input("Detected By", value=row.get('Detected By', row.get('Executed By', '')), key=f"det_by_{d_key}")
                            adm_date_origin = st.text_input("Date of Defect Origin", value=row.get('Date of Defect Origin', row.get('Executed Date', '')), key=f"det_org_{d_key}")

                        adm_def_desc = st.text_area("Defect Description", value=default_desc, height=70, key=f"adm_def_desc_{d_key}")
                        adm_steps = st.text_area("Steps to Reproduce", value=default_steps, height=70, key=f"steps_{d_key}")
                        adm_expected = st.text_area("Expected Results", value=default_expected, height=70, key=f"adm_exp_{d_key}")
                        adm_comments = st.text_input("Comments", value=row.get('Comments', row.get('Remarks', '')), key=f"adm_comm_{d_key}")
                        
                        col_e1, col_e2, col_e3, col_e4 = st.columns(4)
                        with col_e1:
                            adm_exec_by = st.text_input("Executed By", value=row.get('Executed By', ''), key=f"exec_{d_key}")
                        with col_e2:
                            adm_utano = st.text_input("Utano", value=row.get('Utano', ''), key=f"utano_{d_key}")
                        with col_e3:
                            adm_fe = st.text_input("FE", value=row.get('FE', ''), key=f"fe_{d_key}")
                        with col_e4:
                            adm_sibs = st.text_input("SIBS", value=row.get('SIBS', ''), key=f"sibs_{d_key}")

                        st.divider()
                        
                        sev_options = ["Low", "Medium", "High", "Critical"]
                        curr_sev = row.get('Severity', 'Medium')
                        if curr_sev not in sev_options: curr_sev = "Medium"
                        
                        pri_options = ["Low", "Moderate", "High"]
                        curr_pri = str(row.get('Priority', 'Moderate')).strip()
                        if curr_pri in ["Medium", "medium"]:
                            curr_pri = "Moderate"
                        if curr_pri not in pri_options: 
                            curr_pri = "Moderate"
                        
                        stat_options = ["Open", "In Progress", "Resolved", "Closed", "Rejected"]
                        curr_stat = row.get('Defect Status', 'Open')
                        if curr_stat not in stat_options: curr_stat = "Open"
                        
                        assign_options = ["Development Team", "Tester", "Vendor (GRG)", "Network Team"]
                        curr_assign = row.get('Assigned To', 'Development Team')
                        if curr_assign not in assign_options: curr_assign = "Development Team"

                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            adm_severity = st.selectbox("Severity", sev_options, index=sev_options.index(curr_sev), key=f"sev_{d_key}")
                            adm_status = st.selectbox("Defect Status", stat_options, index=stat_options.index(curr_stat), key=f"stat_{d_key}")
                        with col_m2:
                            adm_priority = st.selectbox("Priority", pri_options, index=pri_options.index(curr_pri), key=f"pri_{d_key}")
                            adm_assigned = st.selectbox("Assigned To", assign_options, index=assign_options.index(curr_assign), key=f"assign_{d_key}")
                        with col_m3:
                            today_dt = datetime.today()
                            try:
                                t_val = row.get('Target Date', '')
                                default_dt = datetime.strptime(t_val, "%Y-%m-%d").date() if t_val else today_dt
                            except Exception:
                                default_dt = today_dt
                            adm_target_date = st.date_input("Expected Date of Closure / Target Date", value=default_dt, key=f"tdate_{d_key}").strftime("%Y-%m-%d")

                        curr_root = row.get('Root Cause', '')
                        adm_root_cause = st.text_area("Root Cause / Developer Resolution Notes", value=curr_root, height=80, key=f"root_{d_key}")

                        submitted_admin_def = st.form_submit_button(f"💾 Save All Defect Changes to Database ({tc_id})", use_container_width=True)

                        if submitted_admin_def:
                            admin_update_full_defect_details(
                                tc_id, mod_name, adm_steps, actual_result=adm_def_desc, executed_by=adm_exec_by, utano=adm_utano, 
                                fe=adm_fe, sibs=adm_sibs, severity=adm_severity, priority=adm_priority, defect_status=adm_status, 
                                assigned_to=adm_assigned, target_date=adm_target_date, root_cause=adm_root_cause,
                                defect_desc=adm_def_desc
                            )
                            st.success(f"Defect report updated successfully for {tc_id}!")
                            st.rerun()
                else:
                    c_inf1, c_inf2 = st.columns(2)
                    with c_inf1:
                        st.markdown("**Defect Description:**")
                        st.info(default_desc)
                        st.markdown("**Steps to Reproduce:**")
                        st.warning(default_steps)
                        st.markdown("**Expected Results:**")
                        st.success(default_expected)
                    with c_inf2:
                        st.markdown(f"**Detected By:** `{row.get('Detected By', row.get('Executed By', ''))}` | **Date of Defect Origin:** `{row.get('Date of Defect Origin', row.get('Executed Date', ''))}`")
                        st.markdown(f"**Application / Module:** `{mod_name}` | **Comments:** `{row.get('Comments', row.get('Remarks', ''))}`")
                        st.markdown(f"**Utano:** `{row.get('Utano', '')}` | **RRN:** `{row.get('RRN', '')}` | **FE:** `{row.get('FE', '')}` | **SIBS:** `{row.get('SIBS', '')}`")

                    st.divider()
                    st.markdown("#### 📋 Defect Matrix Details")
                    col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                    col_r1.metric("Severity", row.get('Severity') if row.get('Severity') else "Unassigned")
                    col_r2.metric("Priority", row.get('Priority') if row.get('Priority') else "Unassigned")
                    col_r3.metric("Defect Status", row.get('Defect Status') if row.get('Defect Status') else "Open")
                    col_r4.metric("Assigned To", row.get('Assigned To') if row.get('Assigned To') else "Unassigned")
                    
                    if row.get('Root Cause'):
                        st.info(f"**Root Cause / Resolution Notes:** {row['Root Cause']}")

# ---------------------------------------------------------
# 4. ADMIN MANAGEMENT
# ---------------------------------------------------------
elif menu == "⚙️ Admin Management":
    st.subheader("⚙️ Admin Management & Test Case Creation / Deletion")
    
    is_admin = (st.session_state.authenticated_role == "Admin / Manager")
    
    if not is_admin:
        st.warning("⚠️ You must be logged in as an **Admin / Manager** to manage test cases.")
    else:
        tab_add, tab_del = st.tabs(["➕ Add New Test Case", "🗑️ Search & Delete Test Cases"])
        
        with tab_add:
            st.markdown("### ➕ Add New Test Case")
            
            # Fetch existing unique modules from current data for the dropdown
            existing_modules = list(df['Module Name'].unique()) if not df.empty and 'Module Name' in df.columns else [
                "Card_Based_Cash Deposit-Positive", "CardBased Cash Deposit-Negative", 
                "Credit_Card_Positive_Scenario", "Credit_Card_Negative_Scenario"
            ]
            
            # --- OUTSIDE FORM: Dynamic Module Selector (Updates instantly) ---
            mod_choice_type = st.radio("Module Input Mode", ["Select Existing Module", "Type Custom Module"], key="mod_mode_radio", horizontal=True)
            
            if mod_choice_type == "Select Existing Module":
                new_module = st.selectbox("Select Module Name", existing_modules, key="sel_existing_mod")
            else:
                custom_mod_input = st.text_input("Type Custom Module Name (Exact match)", placeholder="e.g., Other Bank Cards", key="txt_custom_mod")
                new_module = custom_mod_input.strip()

            st.divider()

            # --- INSIDE FORM: Test Case Details ---
            with st.form("new_test_case_form"):
                col_n1, col_n2, col_n3 = st.columns(3)
                with col_n1:
                    new_tc_id = st.text_input("Test Case ID (e.g., TC_CB_CR_P_01_005)")
                with col_n2:
                    new_category = st.selectbox("Category", ["Card Based", "Cardless", "Other Bank Cards"])
                    new_path_type = st.selectbox("Path Type", ["Positive", "Negative", "Edge Case"])
                with col_n3:
                    new_test_area = st.text_input("Test Area (e.g., Credit Card Limit)")
                    
                new_test_desc = st.text_input("Test Case Description")
                new_pre_cond = st.text_area("Pre-Conditions", height=70)
                new_test_steps = st.text_area("Test Steps (Numbered list)", height=100)
                new_exp_result = st.text_area("Expected Result", height=80)
                
                submitted = st.form_submit_button("🚀 Insert New Test Case to Database", use_container_width=True)
                if submitted:
                    if not new_tc_id or not new_module or not new_test_desc:
                        st.error("Please fill in at least Test Case ID, Module Name, and Description.")
                    else:
                        success = insert_new_test_case_to_db(
                            new_tc_id, new_category, new_module, new_test_area, 
                            new_test_desc, new_pre_cond, new_test_steps, new_exp_result, new_path_type
                        )
                        if success:
                            st.success(f"Successfully added test case {new_tc_id} under module '{new_module}'! It will now appear in your dropdowns.")
                            st.rerun()

        with tab_del:
            st.markdown("### 🗑️ Search & Delete Existing Test Cases")
            st.markdown("Use the search options below to find specific test cases and securely remove them from the database.")
            
            del_search = st.text_input("🔍 Search Test Case by ID, Description, or Module", placeholder="e.g. TC_CB_CD_P_01_001")
            
            search_df = df.copy()
            if del_search and not search_df.empty:
                search_df = search_df[
                    search_df['TC ID'].str.contains(del_search, case=False, na=False) | 
                    search_df['Test Case Description'].str.contains(del_search, case=False, na=False) | 
                    search_df['Module Name'].str.contains(del_search, case=False, na=False)
                ]
            
            if search_df.empty:
                st.info("No matching test cases found.")
            else:
                st.write(f"Found {len(search_df)} test case(s):")
                
                for idx, row in search_df.iterrows():
                    tc_id_val = row['TC ID']
                    mod_name_val = row['Module Name']
                    desc_val = row['Test Case Description']
                    
                    with st.expander(f"📌 [{tc_id_val}] {desc_val} (Module: {mod_name_val})"):
                        col_d1, col_d2 = st.columns([3, 1])
                        with col_d1:
                            st.markdown(f"**Category:** `{row.get('Category')}`")
                            st.markdown(f"**Path Type:** `{row.get('Path Type')}`")
                            st.markdown(f"**Current Status:** `{row.get('Status')}`")
                        with col_d2:
                            del_btn_key = f"del_tc_{tc_id_val}_{mod_name_val}_{idx}"
                            if st.button("🗑️ Delete Test Case", key=del_btn_key, type="primary"):
                                deleted = delete_test_case_from_db(tc_id_val, mod_name_val)
                                if deleted:
                                    st.success(f"Successfully deleted test case {tc_id_val}!")
                                    st.rerun()

            st.divider()
            st.markdown("### 🗑️ Delete Entire Module & All Its Test Cases")
            st.markdown("If a custom module was added by mistake, you can wipe the entire module and its associated test cases from the database here:")
            
            all_existing_modules = list(df['Module Name'].dropna().unique()) if not df.empty and 'Module Name' in df.columns else []
            
            if all_existing_modules:
                mod_to_delete = st.selectbox("Select Module Name to Delete", all_existing_modules, key="select_mod_to_del")
                
                # Confirmation checkbox to prevent accidental deletions
                confirm_mod_del = st.checkbox(f"⚠️ Yes, I want to permanently delete module '{mod_to_delete}' and all test cases under it.", key="chk_confirm_mod_del")
                
                if st.button("🗑️ Delete Entire Module", type="primary", key="btn_del_entire_mod", use_container_width=True):
                    if not confirm_mod_del:
                        st.error("Please check the confirmation box above before deleting a module.")
                    else:
                        # Filter rows belonging to this module and delete them one by one or via bulk handler
                        module_rows = df[df['Module Name'] == mod_to_delete]
                        success_count = 0
                        for _, m_row in module_rows.iterrows():
                            if delete_test_case_from_db(m_row['TC ID'], m_row['Module Name']):
                                success_count += 1
                        
                        st.success(f"Successfully deleted module '{mod_to_delete}' along with {success_count} associated test case(s)!")
                        st.rerun()
            else:
                st.info("No modules available for deletion.")

    st.divider()
    st.info("To reload or refresh test cases from your master files, use the administrative database tools configured for your Supabase project.")

# ---------------------------------------------------------
# 5. REPORTS
# ---------------------------------------------------------
elif menu == "📄 Reports":
    st.subheader("📄 Filter & Export Customized UAT Reports")
    
    can_report = st.session_state.authenticated_role in ["Admin / Manager", "Tester"]
    if not can_report:
        st.warning("⚠️ You are logged in as a Viewer role. Advanced custom reporting downloads are restricted to **Testers** and **Admin / Managers**.")
    
    st.markdown("Configure your multi-select criteria and date filters below to generate a tailored report export:")
    
    rep_df = df.copy()
    
    all_cats = list(rep_df['Category'].unique()) if not rep_df.empty and 'Category' in rep_df.columns else []
    all_mods = list(rep_df['Module Name'].unique()) if not rep_df.empty and 'Module Name' in rep_df.columns else []
    all_testers = list(rep_df['Executed By'].dropna().unique()) if not rep_df.empty and 'Executed By' in rep_df.columns else []
    
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        rep_cats = st.multiselect("Select Categories", all_cats, default=all_cats, key="rep_cats")
    with col_f2:
        rep_mods = st.multiselect("Select Modules / Sheets", all_mods, default=all_mods, key="rep_mods")
    with col_f3:
        rep_paths = st.multiselect("Select Path Types", ["Positive", "Negative"], default=["Positive", "Negative"], key="rep_paths")

    col_f4, col_f5 = st.columns(2)
    with col_f4:
        rep_statuses = st.multiselect("Select Statuses", ["PASS", "FAIL", "N/A", "PENDING"], default=["PASS", "FAIL", "N/A", "PENDING"], key="rep_statuses")
    with col_f5:
        rep_testers = st.multiselect("Select Testers (Executed By)", all_testers, default=all_testers, key="rep_testers")

    st.divider()
    st.markdown("#### 📅 Date Filter Configuration")
    d_mode_col1, d_mode_col2 = st.columns(2)
    with d_mode_col1:
        enable_rep_date_filter = st.checkbox("Enable Execution Date Filter for Download", key="rep_date_chk")
    with d_mode_col2:
        if enable_rep_date_filter:
            date_filter_mode = st.radio("Date Filter Mode", ["Date Range", "Specific Date"], horizontal=True, key="rep_date_mode")
        else:
            date_filter_mode = None

    rep_date_range = None
    specific_rep_date = None

    if enable_rep_date_filter:
        if date_filter_mode == "Date Range":
            rep_date_range = st.date_input("Select Execution Date Range", value=(date.today(), date.today()), key="rep_date_rng")
        else:
            specific_rep_date = st.date_input("Select Specific Execution Date", value=date.today(), key="rep_specific_date")

    if not rep_df.empty:
        if 'Category' in rep_df.columns:
            rep_df = rep_df[rep_df['Category'].isin(rep_cats)]
        if 'Module Name' in rep_df.columns:
            rep_df = rep_df[rep_df['Module Name'].isin(rep_mods)]
        if 'Path Type' in rep_df.columns:
            rep_df = rep_df[rep_df['Path Type'].isin(rep_paths)]
        if 'Status' in rep_df.columns:
            rep_df = rep_df[rep_df['Status'].isin(rep_statuses)]
        if 'Executed By' in rep_df.columns and all_testers:
            rep_df = rep_df[rep_df['Executed By'].isin(rep_testers)]

        if enable_rep_date_filter:
            def parse_dt(val):
                if not val or pd.isna(val): return None
                try:
                    return pd.to_datetime(val).date()
                except Exception:
                    return None
            
            exec_dates = rep_df['Executed Date'].apply(parse_dt)
            
            if date_filter_mode == "Date Range" and isinstance(rep_date_range, tuple) and len(rep_date_range) == 2:
                start_d, end_d = rep_date_range
                mask = exec_dates.apply(lambda d: d is not None and start_d <= d <= end_d)
                rep_df = rep_df[mask]
            elif date_filter_mode == "Specific Date" and specific_rep_date:
                mask = exec_dates.apply(lambda d: d is not None and d == specific_rep_date)
                rep_df = rep_df[mask]

    st.write(f"📊 **Filtered Results Preview ({len(rep_df)} test cases match criteria):**")
    
    # Comprehensive columns list ensuring nothing gets cut off
    comprehensive_cols = [
        'TC ID', 'Category', 'Module Name', 'Test Area', 'Test Case Description', 
        'Path Type', 'Pre-Conditions', 'Test Steps', 'Expected Result', 'Actual Result', 
        'Status', 'Executed By', 'Executed Date', 'RRN', 'Utano', 'FE', 'SIBS', 'Remarks'
    ]
    
    for col in comprehensive_cols:
        if col not in rep_df.columns:
            rep_df[col] = ""

    preview_df = rep_df[comprehensive_cols] if not rep_df.empty else pd.DataFrame(columns=comprehensive_cols)
    st.dataframe(preview_df, use_container_width=True, height=350, hide_index=True)

    st.write("")
    report_output_df = rep_df[comprehensive_cols] if not rep_df.empty else pd.DataFrame(columns=comprehensive_cols)
    report_buf = generate_professional_report_excel(report_output_df, report_title="CUSTOMIZED UAT TEST EXECUTION REPORT")
    
    st.download_button(
        label="📥 Download Filtered UAT Report (.xlsx)",
        data=report_buf.getvalue(),
        file_name=f"PeoplesBank_CRM_UAT_Filtered_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        disabled=not can_report
    )

# --- FOOTER ---
st.markdown("""
    <div class="custom-footer">
        PEOPLE'S BANK — GRG CRM UAT PORTAL &nbsp;|&nbsp; Developed by <b>Dhinesh Melroy - IT Graduate Trainee(Channel_Department)</b>
    </div>
""", unsafe_allow_html=True)